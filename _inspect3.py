import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

path = r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm"
wb = load_workbook(path, data_only=False, keep_vba=True)
ws = wb["Calculos"]

# extent of reading formulas / empties
print("=== Instrument block1 C20:D22 / Pattern F20:G22 ===")
for r in range(20, 23):
    print(f"C{r}={ws[f'C{r}'].value!r} D{r}={ws[f'D{r}'].value!r} F{r}={ws[f'F{r}'].value!r} G{r}={ws[f'G{r}'].value!r}")

print("\n=== Main instrument C28:F40 (type) ===")
for r in range(28, 41):
    kinds = []
    for col in "CDEF":
        v = ws[f"{col}{r}"].value
        kinds.append("F" if isinstance(v,str) and v.startswith("=") else ("N" if v is not None else "."))
    print(f"row{r} CDEF={''.join(kinds)} A={repr(ws[f'A{r}'].value)[:60] if ws[f'A{r}'].value else None}")

print("\n=== Main pattern H28:K40 ===")
for r in range(28, 41):
    kinds = []
    for col in "HIJK":
        v = ws[f"{col}{r}"].value
        kinds.append("F" if isinstance(v,str) and v.startswith("=") else ("N" if isinstance(v,(int,float)) else ("S" if isinstance(v,str) else ".")))
    print(f"row{r} HIJK={''.join(kinds)} vals={[ws[f'{c}{r}'].value for c in 'HIJK']}")

print("\n=== EMP Q28:Q40 ===")
for r in range(28, 41):
    print(f"Q{r}={ws[f'Q{r}'].value!r} locked={ws[f'Q{r}'].protection.locked}")

print("\n=== J11 formula ===")
print(repr(ws["J11"].value))
print("J13=", repr(ws["J13"].value)[:250])

# All sheets protection via openpyxl already done
# Try win32com briefly
try:
    import win32com.client
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    wb2 = excel.Workbooks.Open(path, ReadOnly=True)
    sh = wb2.Worksheets("Calculos")
    print("\n=== win32com Calculos ProtectContents ===", sh.ProtectContents)
    print("ProtectDrawingObjects", sh.ProtectDrawingObjects)
    print("ProtectionMode", sh.ProtectionMode)
    wb2.Close(False)
    excel.Quit()
    print("win32com OK")
except Exception as e:
    print("win32com skip/err:", type(e).__name__, e)
