import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook

path = r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm"
wb = load_workbook(path, data_only=False, keep_vba=True)
ws = wb["Calculos"]

for coord in ["K4","K5","K6","K7","K8","L8","F15","G15","A15","C14","J9","J10","F10","B10","B11","B12","Q28","H28","C28"]:
    c = ws[coord]
    print(f"{coord}: value={c.value!r} locked={c.protection.locked}")

print("\n--- K4:K8 / L ---")
for r in range(4, 9):
    print(f"K{r}={ws[f'K{r}'].value!r} L{r}={ws[f'L{r}'].value!r}")

print("\n--- row 4-8 ---")
for r in range(4, 9):
    for c in range(1, 15):
        cell = ws.cell(r, c)
        if cell.value is not None:
            print(f"{cell.coordinate}={cell.value!r}")

print("\n--- pattern H28:K38 ---")
for r in range(28, 39):
    vals = [ws.cell(r, c).value for c in range(8, 12)]
    print(f"H{r}:K{r} = {vals}")

print("\n--- lock flags ---")
for coord in ["J9","F10","J10","F15","B13","F13","F14","J14","C28","H28","Q28","K4","D4","E4","F4","B10","B11","B12"]:
    c = ws[coord]
    print(f"{coord}: locked={c.protection.locked} val={repr(c.value)[:100]}")

print("L8=", ws["L8"].value)

# Merges intersecting key inputs
inputs = ["B10","B11","B12","J9","F10","J10","C14","F15","C28","H28","Q28"]
from openpyxl.utils import range_boundaries, coordinate_to_tuple
print("\n--- merges intersecting inputs ---")
for m in ws.merged_cells.ranges:
    min_col, min_row, max_col, max_row = range_boundaries(str(m))
    for inp in inputs + ["A10","A11","A12","H9","E10","H10","A14","A15","A18","F18","A24","H24","Q25"] + [f"C{r}" for r in range(28,39)] + [f"H{r}" for r in range(28,39)] + [f"Q{r}" for r in range(28,39)] + ["F15","G15","B9","B10","B11","B12","F10","G10","J9","J10"]:
        # check if inp in merge
        try:
            r, c = coordinate_to_tuple(inp) if False else None
        except: pass
    # simpler: check each merge against known ranges of interest
    interest = [
        ("B10",10,2),("C10",10,3),("B11",11,2),("C11",11,3),("B12",12,2),("C12",12,3),
        ("J9",9,10),("F10",10,6),("G10",10,7),("J10",10,10),("C14",14,3),("A14",14,1),("B14",14,2),
        ("A15",15,1),("F15",15,6),("G15",15,7),("H9",9,8),("I9",9,9),("H10",10,8),("I10",10,9),
        ("Q28",28,17),("C28",28,3),("H28",28,8),("A24",24,1),("H24",24,8),("A18",18,1),("F18",18,6),
        ("Q25",25,17),("F11",11,6),
    ]
    for name, rr, cc in interest:
        if min_row <= rr <= max_row and min_col <= cc <= max_col:
            print(f"  {m} intersects {name}")
