# -*- coding: utf-8 -*-
"""Read-only inspection of Formato master .xlsm"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

PATH = r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm"
OUT = r"C:\Users\AG\Desktop\AGG\project\_inspect_calculos.txt"

wb = load_workbook(PATH, data_only=False, keep_vba=True)
lines = []
def w(s=""):
    lines.append(str(s))

w("=== SHEETS ===")
for s in wb.sheetnames:
    ws = wb[s]
    w(f"  {s!r}  sheet_state={ws.sheet_state}  dimensions={ws.dimensions}")

calc_name = None
for n in wb.sheetnames:
    if "calcul" in n.lower():
        calc_name = n
        break
if not calc_name:
    calc_name = wb.sheetnames[0]
w(f"\n=== USING SHEET: {calc_name!r} ===")
ws = wb[calc_name]

w("\n=== MERGED RANGES (Calculos) ===")
for mr in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col)):
    w(f"  {mr}")

w("\n=== DUMP A1:AE45 nonempty (value/formula + number_format) ===")
for row in range(1, 46):
    for col in range(1, 31):
        cell = ws.cell(row=row, column=col)
        v = cell.value
        if v is None or v == "":
            continue
        addr = f"{get_column_letter(col)}{row}"
        kind = "FORMULA" if isinstance(v, str) and v.startswith("=") else "CONST"
        nf = cell.number_format
        locked = cell.protection.locked
        hidden = cell.protection.hidden
        vs = repr(v)
        if len(vs) > 220:
            vs = vs[:220] + "..."
        w(f"  {addr}\t{kind}\tlocked={locked}\thidden={hidden}\tnf={nf!r}\t{vs}")

w("\n=== DATA VALIDATIONS ===")
dvs = list(ws.data_validations.dataValidation) if ws.data_validations else []
for i, dv in enumerate(dvs):
    w(f"  DV[{i}]: type={dv.type!r} operator={dv.operator!r} allow_blank={dv.allow_blank}")
    w(f"    sqref={dv.sqref}")
    w(f"    formula1={dv.formula1!r}")
    w(f"    formula2={dv.formula2!r}")
    w(f"    showError={dv.showErrorMessage} errorTitle={dv.errorTitle!r} error={dv.error!r}")
    w(f"    showInput={dv.showInputMessage} promptTitle={dv.promptTitle!r} prompt={dv.prompt!r}")

w("\n=== SHEET PROTECTION (openpyxl) ===")
sp = ws.protection
attrs = [
    "sheet","password","formatCells","formatColumns","formatRows","insertColumns",
    "insertRows","insertHyperlinks","deleteColumns","deleteRows","selectLockedCells",
    "selectUnlockedCells","sort","autoFilter","pivotTables","objects","scenarios",
]
for a in attrs:
    w(f"  {a}={getattr(sp, a, None)!r}")

w("\n=== LOCKED/UNLOCKED SUMMARY A1:AE45 ===")
unlocked = []
for row in range(1, 46):
    for col in range(1, 31):
        cell = ws.cell(row=row, column=col)
        if cell.protection.locked is False:
            unlocked.append(f"{get_column_letter(col)}{row}")
w(f"  unlocked cells ({len(unlocked)}): {', '.join(unlocked) if unlocked else '(none)'}")

w("\n=== DEFINED NAMES ===")
try:
    for name in wb.defined_names:
        defn = wb.defined_names[name]
        w(f"  {name}: {defn.attr_text}")
except Exception as e:
    w(f"  error: {e}")

keywords = ["lugar", "sitio", "laboratorio", "emp", "marca", "modelo", "serie", "alcance", "divisi", "unidad", "ibc", "patr", "acuerdo", "en sitio", "norma", "fabricante", "cliente", "analog", "digital"]
w("\n=== OTHER SHEETS QUICK LABEL SCAN ===")
for sn in wb.sheetnames:
    s = wb[sn]
    hits = []
    max_r = min(s.max_row or 1, 100)
    max_c = min(s.max_column or 1, 50)
    for row in range(1, max_r + 1):
        for col in range(1, max_c + 1):
            v = s.cell(row=row, column=col).value
            if isinstance(v, str) and any(k in v.lower() for k in keywords):
                hits.append(f"{get_column_letter(col)}{row}={v[:100]!r}")
    if hits:
        w(f"  [{sn}] hits={len(hits)}")
        for h in hits[:60]:
            w(f"    {h}")
        if len(hits) > 60:
            w(f"    ... +{len(hits)-60} more")

# Historial dump if exists
for sn in wb.sheetnames:
    if "historial" in sn.lower() or "histor" in sn.lower():
        s = wb[sn]
        w(f"\n=== HISTORIAL SHEET DUMP ({sn}) first 30 rows x 20 cols ===")
        for row in range(1, 31):
            for col in range(1, 21):
                cell = s.cell(row=row, column=col)
                v = cell.value
                if v is None or v == "":
                    continue
                addr = f"{get_column_letter(col)}{row}"
                vs = repr(v)
                if len(vs) > 120:
                    vs = vs[:120] + "..."
                w(f"  {addr}\t{vs}")

# Focus cells around header area for Lugar/EMP etc - also dump formulas nearby
w("\n=== FOCUS: row1-15 cols A-P values ===")
for row in range(1, 16):
    rowvals = []
    for col in range(1, 17):
        v = ws.cell(row=row, column=col).value
        if v is not None and v != "":
            rowvals.append(f"{get_column_letter(col)}={v!r}"[:80])
    if rowvals:
        w(f"  R{row}: " + " | ".join(rowvals))

with open(OUT, "w", encoding="utf-8") as f:
    f.write("\n".join(lines))
print(f"Wrote {OUT} lines={len(lines)}")
print("Sheets:", wb.sheetnames)
print("Calculos:", calc_name)
