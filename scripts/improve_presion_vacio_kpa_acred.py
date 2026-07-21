#!/usr/bin/env python3
"""
Patron por alcance + vacio/kPa segun acreditacion PJLA L25-682:

  Positivo:
    0 .. 750 psi   (0 .. 5171.07 kPa)  → AG-052
    0 .. 10000 psi (0 .. 68947.59 kPa) → AG-008
    (hasta 350 psi tambien puede AG-034)

  Negativo / vacio (acreditado):
    -11.3 .. 350.5 psi  (−77.9 .. 2416.5 kPa) → AG-034 Fluke PV350
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_clase.xlsm")
ALT = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion.xlsm")

# Alcance F10 → psi (soporta psi / kPa / bar)
F_ALC_PSI = (
    '=IF(F10="","",'
    'IF(OR(J10="",J10="psi",J10="PSI"),F10,'
    'IF(OR(J10="kPa",J10="KPA",J10="kpa"),F10/6.894757293,'
    'IF(OR(J10="bar",J10="BAR",J10="Bar"),F10/0.06894757293,F10))))'
)

# Rango minimo N13 → psi (vacio si < 0)
F_MIN_PSI = (
    '=IF(OR(N13="",N13=0),0,'
    'IF(OR(J10="",J10="psi",J10="PSI"),N13,'
    'IF(OR(J10="kPa",J10="KPA",J10="kpa"),N13/6.894757293,'
    'IF(OR(J10="bar",J10="BAR",J10="Bar"),N13/0.06894757293,N13))))'
)

# Seleccion patron (vacio → siempre AG-034)
F_ID = (
    '=IF(O13="","",'
    'IF(P13<0,"AG-034",'
    'IF(O13<=350,"AG-034",'
    'IF(O13<=750,"AG-052","AG-008"))))'
)

F_NOM = (
    '=IF(N12="","",'
    'IF(P13<0,"Fluke PV350 VACIO (acred. -11.3 a 350.5 psi / -77.9 a 2417 kPa)",'
    'IF(N12="AG-034","Fluke PV350 (0-350 psi / 0-2414 kPa)",'
    'IF(N12="AG-052","0-750 psi / 0-5171 kPa (AG-052)",'
    '"Omega 0-10000 psi / 0-68948 kPa (AG-008)"))))'
)

# Aviso si se sale del alcance acreditado
F_WARN = (
    '=IF(O13="","",'
    'IF(AND(P13<0,P13<-11.3),"ALERTA: vacio bajo -11.3 psi (-77.9 kPa) fuera de acreditacion PJLA",'
    'IF(AND(P13<0,O13>350.5),"ALERTA: vacio/compuesto sobre 350.5 psi fuera de acreditacion PJLA",'
    'IF(AND(P13>=0,O13>10000),"ALERTA: alcance > 10000 psi fuera de acreditacion PJLA",'
    'IF(AND(P13>=0,O13>750,O13<=10000),"","")'
    '))))'
)


def main() -> int:
    path = PATH if PATH.exists() else ALT
    print(f"Archivo: {path}")
    try:
        wb = openpyxl.load_workbook(path, keep_vba=True)
    except PermissionError:
        print("Cierra el Excel e intenta de nuevo.")
        return 1

    calc = wb["Calculos"]
    pat = wb["Patrones"]

    calc["L12"].value = "Patron (acred. PJLA):"
    calc["L12"].font = Font(bold=True, size=9, color="1F4E79")

    calc["L13"].value = "Rango min (vacio):"
    calc["L13"].font = Font(size=9, color="666666")
    # N13: entrada manual; si vacio = 0 (solo positivo)
    if not isinstance(calc["N13"].value, (int, float)):
        calc["N13"].value = None

    calc["O13"].value = F_ALC_PSI  # alcance en psi
    calc["P13"].value = F_MIN_PSI  # minimo en psi
    calc["O12"].value = "alcance→psi"
    calc["P12"].value = "min→psi"
    calc["O12"].font = Font(size=8, italic=True, color="888888")
    calc["P12"].font = Font(size=8, italic=True, color="888888")

    calc["N12"].value = F_ID
    calc["N12"].font = Font(bold=True, color="006600")
    calc["M12"].value = F_NOM
    calc["M12"].font = Font(italic=True, size=9)

    calc["L14"].value = "Aviso acreditacion:"
    calc["L14"].font = Font(size=9, color="C00000")
    calc["M14"].value = F_WARN
    calc["M14"].font = Font(bold=True, color="C00000", size=9)

    # Tabla acreditacion en Patrones
    pat["A36"].value = "Seleccion automatica + alcance acreditado PJLA L25-682 (Presion)"
    pat["A36"].font = Font(bold=True, color="1F4E79")

    headers = ["Condicion", "ID", "Equipo", "Alcance acreditado", "En kPa (equiv.)"]
    for i, h in enumerate(headers, 1):
        cell = pat.cell(37, i, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")

    data = [
        ("Vacio / negativo (N13<0)", "AG-034", "Fluke PV350", "-11.3 a 350.5 psi", "-77.9 a 2416.5 kPa"),
        ("Positivo <= 350 psi", "AG-034", "Fluke PV350", "0 a 350 psi (uso lab)", "0 a 2414 kPa"),
        ("Positivo <= 750 psi", "AG-052", "Druck DPI104 / Omega scope", "0 a 750 psi", "0 a 5171.07 kPa"),
        ("Positivo > 750 psi", "AG-008", "Omega DPG-4000-10K", "0 a 10000 psi", "0 a 68947.59 kPa"),
    ]
    for r, row in enumerate(data, 38):
        for c, val in enumerate(row, 1):
            pat.cell(r, c, val)

    pat["A42"].value = (
        "Como usar vacio en kPa: J10=kPa, F10=alcance max (ej. 2417), N13=min negativo (ej. -77.9). "
        "Patron → AG-034. Si N13 < -77.9 kPa aparece alerta fuera de acreditacion."
    )
    pat["A42"].font = Font(italic=True, size=9)

    try:
        wb.save(path)
        print(f"Guardado: {path}")
    except PermissionError:
        alt = path.with_name(path.stem + "_vacio" + path.suffix)
        wb.save(alt)
        print(f"Abierto. Guardado como: {alt}")

    print("OK: vacio/kPa alineado a PJLA (-11.3..350.5 psi → AG-034)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
