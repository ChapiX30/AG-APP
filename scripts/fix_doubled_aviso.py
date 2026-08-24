# -*- coding: utf-8 -*-
from __future__ import annotations

import re
import shutil
import sys
import zipfile
from io import BytesIO
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

FOLDER = Path(r"C:\Users\AG\Desktop\FORMATOS AG")
MSG = "⚠ PATRÓN VENCIDO — Actualiza vigencia o cambia el patrón"
PAT_PROT = re.compile(
    rb"<sheetProtection\b[^>]*/>|<sheetProtection\b[\s\S]*?</sheetProtection>", re.I
)


def strip_protection(src: Path, dst: Path) -> None:
    with zipfile.ZipFile(src, "r") as zin:
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                    data, _ = PAT_PROT.subn(b"", data)
                zout.writestr(info, data)
        dst.write_bytes(buf.getvalue())


def needs_fix(path: Path) -> bool:
    wb = openpyxl.load_workbook(path, keep_vba=True, read_only=True)
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(max_row=min(ws.max_row or 1, 90), max_col=min(ws.max_column or 1, 20)):
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and (
                        "⚠ ⚠" in v or "patrón — Actualiza" in v.lower() or "patrón — actualiza" in v.lower()
                    ):
                        return True
    finally:
        wb.close()
    return False


def fix_formula(v: str | None) -> str | None:
    if not isinstance(v, str) or not v.startswith("="):
        return None
    if "⚠ ⚠" not in v and "— Actualiza vigencia o cambia el patrón — Actualiza" not in v:
        return None

    def repl(m: re.Match) -> str:
        inner = m.group(1)
        if "Actualiza vigencia" in inner or "⚠" in inner:
            return f'"{MSG}"'
        return m.group(0)

    return re.sub(r'"([^"]*)"', repl, v)


def main() -> int:
    fill = PatternFill("solid", fgColor="FFDCDC")
    font = Font(name="Calibri", bold=True, size=10, color="991B1B")
    align = Alignment(wrap_text=True, vertical="center")

    for path in sorted(FOLDER.glob("Formato*.xlsm")):
        if not needs_fix(path):
            continue
        work = path.with_name("_fix_aviso_tmp.xlsm")
        strip_protection(path, work)
        wb = openpyxl.load_workbook(work, keep_vba=True)
        n = 0
        for ws in wb.worksheets:
            if ws.title.startswith("BD_"):
                continue
            for row in ws.iter_rows(max_row=min(ws.max_row or 1, 90), max_col=min(ws.max_column or 1, 20)):
                for cell in row:
                    raw = cell.value if isinstance(cell.value, str) else None
                    new = fix_formula(raw)
                    if new and new != raw:
                        cell.value = new
                        cell.fill = fill
                        cell.font = font
                        cell.alignment = align
                        n += 1
        wb.save(work)
        wb.close()
        shutil.copy2(work, path)
        work.unlink(missing_ok=True)
        print("fixed", path.name, n)
    print("done")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
