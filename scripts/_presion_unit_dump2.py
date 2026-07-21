import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path

path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
ws = wb["Calculos"]

print("=== Row 27 headers A-AF ===")
for c in range(1, 32):
    v = ws.cell(row=27, column=c).value
    if v: print(f"  {get_column_letter(c)}27: {v}")

print("\n=== Q column rows 20-30 ===")
for r in range(20, 31):
    v = ws.cell(row=r, column=17).value
    print(f"  Q{r}: {v}")

print("\n=== AD column 28-30 ===")
for r in [28,29,30]:
    v = ws.cell(row=r, column=30).value
    print(f"  AD{r}: {v}")

print("\n=== A28, G28, H28 formulas ===")
for coord in ["A28","G28","H28","P28","R28","S28"]:
    from openpyxl.utils import coordinate_from_string, column_index_from_string
    col, row = coordinate_from_string(coord)
    c = column_index_from_string(col)
    print(f"  {coord}: {ws.cell(row=row,column=c).value}")

print("\n=== J10 data validation ===")
if ws.data_validations and ws.data_validations.dataValidation:
    for dv in ws.data_validations.dataValidation:
        sq = getattr(dv, 'sqref', None) or getattr(dv, 'cells', '')
        if 'J10' in str(sq) or 'J10' in str(dv):
            print(f"  DV sqref={sq} type={dv.type} formula1={dv.formula1} formula2={getattr(dv,'formula2',None)}")
    # print all DVs touching col J row 10
    for dv in ws.data_validations.dataValidation:
        sq = str(getattr(dv, 'sqref', ''))
        if 'J10' in sq:
            print(f"  J10 match: sqref={sq} f1={dv.formula1}")

print("\n=== All data validations on Calculos (formula1) ===")
for dv in ws.data_validations.dataValidation:
    print(f"  {dv.sqref} | {dv.type} | {dv.formula1}")

# Patrones
wp = wb["Patrones"]
print("\n=== Patrones rows 1-20 cols A-I ===")
for r in range(1, 21):
    parts = []
    for c in range(1, 10):
        v = wp.cell(row=r, column=c).value
        if v is not None:
            parts.append(f"{get_column_letter(c)}={str(v)[:80]}")
    if parts: print(f"R{r}: " + " | ".join(parts))

print("\n=== Patrones F6:I16 and headers row 5-6 ===")
for r in range(4, 18):
    parts = [f"{get_column_letter(c)}{r}={wp.cell(row=r,column=c).value}" for c in range(6,10)]
    print("  " + " | ".join(parts))

print("\n=== Patrones row 30+ (AG-034 block) ===")
for r in range(30, 45):
    parts = []
    for c in range(1, 10):
        v = wp.cell(row=r, column=c).value
        if v is not None: parts.append(f"{get_column_letter(c)}={v}")
    if parts: print(f"R{r}: " + " | ".join(parts))

print("\n=== Patrones A45+ ===")
for r in range(45, 70):
    parts = []
    for c in range(1, 12):
        v = wp.cell(row=r, column=c).value
        if v is not None: parts.append(f"{get_column_letter(c)}={str(v)[:100]}")
    if parts: print(f"R{r}: " + " | ".join(parts))

# Search whole wb for conversion
print("\n=== Any sheet cell containing 6.894 or Conversiones ===")
for sn in wb.sheetnames:
    w = wb[sn]
    for row in w.iter_rows(max_row=150):
        for cell in row:
            v = cell.value
            if v is None: continue
            s = str(v)
            if "6.894" in s or "Conversiones" in s or s in ("kPa","MPa","bar","psi"):
                if "6.894" in s or "Conversiones" in s or (cell.coordinate in ("J10",) and sn=="Calculos"):
                    pass
            if "6.894" in s or "145.038" in s or "Conversiones" in s:
                print(f"  {sn}!{cell.coordinate}: {s[:120]}")

wb.close()
