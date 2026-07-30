# -*- coding: utf-8 -*-
"""Comparar micrometro mm vs in."""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

FOLDER = Path(r"C:\Users\AG\Desktop\FORMATOS AG")
MM = FOLDER / "Formato Micrometro Exteriores mm .xlsx"
IN = FOLDER / "Formato Micrometro Exteriores in.xlsx"


def main() -> None:
    for f in (MM, IN):
        print("=" * 60, f.name, "exists", f.exists(), "size", f.stat().st_size if f.exists() else 0)
        wb = openpyxl.load_workbook(f, data_only=False)
        print("sheets:", wb.sheetnames)
        for s in wb.sheetnames:
            ws = wb[s]
            if hasattr(ws, "max_row"):
                print(f"  [{s}] {ws.max_row}x{ws.max_column}")
        wb.close()

    mm = openpyxl.load_workbook(MM, data_only=False)
    inn = openpyxl.load_workbook(IN, data_only=False)

    for sheet in mm.sheetnames:
        if sheet not in inn.sheetnames:
            print(f"Sheet solo en mm: {sheet}")
            continue
        ws1, ws2 = mm[sheet], inn[sheet]
        diffs = []
        max_r = max(ws1.max_row or 1, ws2.max_row or 1)
        max_c = max(ws1.max_column or 1, ws2.max_column or 1)
        for r in range(1, min(max_r, 200) + 1):
            for c in range(1, min(max_c, 40) + 1):
                v1, v2 = ws1.cell(r, c).value, ws2.cell(r, c).value
                if v1 != v2:
                    diffs.append((get_column_letter(c) + str(r), v1, v2))
        print(f"\n=== {sheet}: {len(diffs)} diffs (cap 60) ===")
        for addr, v1, v2 in diffs[:60]:
            print(f"  {addr}: MM={repr(v1)[:100]}")
            print(f"       IN={repr(v2)[:100]}")

    print("\n=== CALCULOS header mm ===")
    ws = mm["CALCULOS"] if "CALCULOS" in mm.sheetnames else mm[mm.sheetnames[2]]
    # find calc sheet
    for name in mm.sheetnames:
        if "CALC" in name.upper():
            ws = mm[name]
            print("using", name)
            break
    for r in range(1, 25):
        row = []
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if v is not None:
                row.append(f"{get_column_letter(c)}{r}={repr(v)[:55]}")
        if row:
            print(f"R{r}:", " | ".join(row))

    print("\n=== CALCULOS header in ===")
    for name in inn.sheetnames:
        if "CALC" in name.upper():
            ws = inn[name]
            print("using", name)
            break
    for r in range(1, 25):
        row = []
        for c in range(1, 20):
            v = ws.cell(r, c).value
            if v is not None:
                row.append(f"{get_column_letter(c)}{r}={repr(v)[:55]}")
        if row:
            print(f"R{r}:", " | ".join(row))

    mm.close()
    inn.close()


if __name__ == "__main__":
    main()
