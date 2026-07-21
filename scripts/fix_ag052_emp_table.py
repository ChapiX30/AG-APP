"""Fix AG-052 EMP table gaps on Patrones."""
import sys
from pathlib import Path
import openpyxl

SRC = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
OUT_FALLBACK = SRC.with_name("Formato master auto Presion_SIN_REF_emp.xlsm")
AG052_DATA = [
    (0, 0, 0.058), (100, 0.044, 0.082), (200, 0.093, 0.058), (300, 0.092, 0.059),
    (400, 0.191, 0.059), (500, 0.199, 0.060), (600, 0.198, 0.061), (700, 0.293, 0.063),
    (800, 0.293, 0.064), (900, 0.297, 0.072), (1000, 0.323, 0.089),
]
VLOOKUP_RANGE = "Patrones!" + chr(36) + "F" + chr(36) + "44:" + chr(36) + "I" + chr(36) + "54"

def inspect_merges(ws):
    return [str(mc) for mc in sorted(ws.merged_cells.ranges, key=lambda r: (r.min_row, r.min_col))
            if mc.max_row >= 43 and mc.min_row <= 55 and mc.max_col >= 5 and mc.min_col <= 9]

def unmerge_blocking(ws):
    to_remove = [str(mc) for mc in list(ws.merged_cells.ranges)
                 if mc.max_row >= 43 and mc.min_row <= 55 and mc.max_col >= 5 and mc.min_col <= 9]
    for coord in to_remove:
        ws.unmerge_cells(coord)
    return to_remove

def write_ag052_table(ws):
    ws.cell(43, 5).value = "AG-052 Druck 1-23207"
    ws.cell(43, 6).value = "Nominal"
    ws.cell(43, 7).value = "Sesgo"
    ws.cell(43, 8).value = "Incert. U(k=2)"
    ws.cell(43, 9).value = "EMP\u00b1"
    for i, (nominal, sesgo, u) in enumerate(AG052_DATA):
        r = 44 + i
        ws.cell(r, 5).value = '=$D$20&"-"&TEXT(F%d,"0")' % r
        ws.cell(r, 6).value = nominal
        ws.cell(r, 7).value = sesgo
        ws.cell(r, 8).value = u
        ws.cell(r, 9).value = abs(sesgo) + u

def verify_vlookup_formulas(wb):
    ws = wb["Calculos"]
    mismatches, ok = [], 0
    for r in range(28, 39):
        for col in (22, 25):
            v = ws.cell(r, col).value
            if isinstance(v, str) and "AG-052" in v:
                if VLOOKUP_RANGE not in v:
                    mismatches.append((r, col, v[:120]))
                else:
                    ok += 1
    return ok, mismatches

def dump_f43_i68(ws):
    lines = ["=== DUMP Patrones F43:I68 ==="]
    for r in range(43, 69):
        lines.append("R%d\tE=%s\tF=%s\tG=%s\tH=%s\tI=%s" % (
            r, ws.cell(r,5).value, ws.cell(r,6).value, ws.cell(r,7).value, ws.cell(r,8).value, ws.cell(r,9).value))
    return "\n".join(lines)

wb = openpyxl.load_workbook(SRC, keep_vba=True, data_only=False)
ws = wb["Patrones"]
print("Before merges:", inspect_merges(ws))
print("Unmerged:", unmerge_blocking(ws))
print("After merges:", inspect_merges(ws))
write_ag052_table(ws)
ok, bad = verify_vlookup_formulas(wb)
print("VLOOKUP ok count:", ok, "range:", VLOOKUP_RANGE)
if bad:
    print("MISMATCHES", bad)
dump = dump_f43_i68(ws)
print(dump)
try:
    wb.save(SRC)
    print("Saved:", SRC)
except PermissionError:
    wb.save(OUT_FALLBACK)
    print("Saved:", OUT_FALLBACK)
Path(r"C:\Users\AG\Desktop\AGG\project\scripts\_ag052_emp_dump.txt").write_text(dump + "\n", encoding="utf-8")
wb.close()
