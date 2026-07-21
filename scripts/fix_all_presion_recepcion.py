#!/usr/bin/env python3
"""Aplica fix Fecha de Recepcion a TODOS los masters Presion (el usuario abre el .xlsm original)."""
from __future__ import annotations

import json
import sys
import urllib.request
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

FORMATOS = Path(r"C:\Users\AG\Desktop\FORMATOS AG")
API = (
    "https://us-central1-agg1-b7f40.cloudfunctions.net/obtenerDatosExcel"
    "?key=TU_CLAVE_SECRETA_AG_APP_2026&prefijo=AGP"
)

HIST_HEADERS = [
    "Name",
    "certificado",
    "cliente",
    "equipo",
    "marca",
    "modelo",
    "serie",
    "id",
    "fecha",
    "tecnico",
    "lugarCalibracion",
    "frecuenciaCalibracion",
    "fechaRecepcion",
]

F_LUGAR = (
    '=IFERROR(VLOOKUP($D$4 & "-" & TEXT($E$4,"0000") & "-" & $F$4,'
    'AG_Historial[[certificado]:[lugarCalibracion]],10,FALSE),"")'
)

F_RECEPCION = (
    '=IFERROR('
    'IF(VLOOKUP($D$4 & "-" & TEXT($E$4,"0000") & "-" & $F$4,'
    'AG_Historial[[certificado]:[fechaRecepcion]],12,FALSE)="",'
    'IF(OR(UPPER(LEFT(K4,1))="S"),"Servicio en Sitio",""),'
    'VALUE(VLOOKUP($D$4 & "-" & TEXT($E$4,"0000") & "-" & $F$4,'
    'AG_Historial[[certificado]:[fechaRecepcion]],12,FALSE))),'
    'IF(OR(UPPER(LEFT(K4,1))="S"),"Servicio en Sitio",""))'
)

F_C14 = (
    '=IF(OR(K4="Laboratorio",K4="laboratorio"),"Instalaciones AG",'
    'IF(OR(K4="Sitio",K4="sitio"),"Instalaciones de Cliente",""))'
)


def fetch_historial():
    req = urllib.request.Request(API, headers={"User-Agent": "AG-FixRecepcion/1.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    return data.get("historial") or []


def find_hist_sheet(wb):
    for name in wb.sheetnames:
        if "obtener" in name.lower():
            return wb[name]
    return wb.create_sheet("obtenerDatosExcel")


def write_historial(ws, rows):
    for col, h in enumerate(HIST_HEADERS, 1):
        ws.cell(1, col, h)
    if ws.max_row and ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for r_i, row in enumerate(rows, 2):
        for c_i, key in enumerate(HIST_HEADERS, 1):
            ws.cell(r_i, c_i, row.get(key, ""))
    n = max(len(rows) + 1, 2)
    last = chr(ord("A") + len(HIST_HEADERS) - 1)
    return f"A1:{last}{n}"


def ensure_table(ws, ref: str):
    for t in list(ws.tables.values()):
        if t.name == "AG_Historial":
            t.ref = ref
            return
    table = Table(displayName="AG_Historial", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(table)


def fix_calculos(calc):
    calc["H4"].value = "Fecha de Recepción:"
    calc["I4"].value = F_RECEPCION
    calc["K4"].value = F_LUGAR
    calc["L4"].value = "<- Lugar (Sitio/Lab)"
    calc["L4"].font = Font(italic=True, size=8, color="888888")
    calc["C14"].value = F_C14


def main() -> int:
    files = sorted(FORMATOS.glob("Formato master auto Presion*.xlsm"))
    if not files:
        print("No hay masters Presion")
        return 1

    print("Descargando historial con fechaRecepcion...")
    historial = fetch_historial()
    print(f"  filas: {len(historial)}")

    for path in files:
        print(f"\n=== {path.name} ===")
        try:
            wb = openpyxl.load_workbook(path, keep_vba=True)
        except PermissionError:
            print("  ABIERTO: cierralo y vuelve a correr este script")
            continue

        hist_ws = find_hist_sheet(wb)
        ref = write_historial(hist_ws, historial)
        try:
            ensure_table(hist_ws, ref)
        except Exception as e:
            print(f"  aviso tabla: {e}")

        fix_calculos(wb["Calculos"])

        # verificar formula
        print("  I4 starts:", str(wb["Calculos"]["I4"].value)[:50])
        print("  header13:", hist_ws.cell(1, 13).value)

        try:
            wb.save(path)
            print("  Guardado OK")
        except PermissionError:
            alt = path.with_name(path.stem + "_FIX_RECEPCION" + path.suffix)
            wb.save(alt)
            print(f"  Estaba bloqueado -> {alt.name}")

    print("\nCierra Excel, abre de nuevo Formato master auto Presion.xlsm y Actualiza datos.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
