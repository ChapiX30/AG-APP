import openpyxl
import re
from pathlib import Path
from openpyxl.utils import get_column_letter

path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(path, data_only=True, keep_vba=True)

def refs(formula):
    if not isinstance(formula, str) or not formula.startswith("="):
        return ""
    fu = formula.upper()
    bits = []
    if "HISTORIAL" in fu: bits.append("Historial")
    for fn in ["INDEX","MATCH","VLOOKUP","XLOOKUP"]:
        if fn in fu: bits.append(fn)
    cols = re.findall(r"Historial!\$([A-Z]+):\$?\1", formula, re.I)
    if cols: bits.append("cols=" + ",".join(sorted(set(cols))))
    return "; ".join(bits)

# Historial headers row 1
if "Historial" in wb.sheetnames:
    ws = wb["Historial"]
    print("Historial row 1 headers:")
    for c in range(1, 20):
        v = ws.cell(1,c).value
        if v: print(f"  {get_column_letter(c)}1: {v!r}")

for sn in ["Calculos", "Portada"]:
    ws, wv = wb[sn], wb_v[sn]
    print(f"\n=== {sn} rows 8-15 ===")
    for r in range(8, 16):
        rowbits = []
        for c in range(1, 20):
            cell = ws.cell(r,c)
            if cell.value is None: continue
            addr = cell.coordinate
            cv = wv[addr].value
            kind = "F" if isinstance(cell.value,str) and cell.value.startswith("=") else "C"
            rowbits.append(f"{addr}({kind})={cell.value!r}/disp={cv!r}")
        if rowbits:
            print(" R"+str(r)+": " + " | ".join(rowbits))

# All formulas on Calculos referencing Historial
print("\n=== Calculos formulas referencing Historial ===")
ws, wv = wb["Calculos"], wb_v["Calculos"]
for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=30):
    for cell in row:
        v = cell.value
        if isinstance(v,str) and v.startswith("=") and "historial" in v.lower():
            print(f"{cell.coordinate}: {v}")
            print(f"  -> {refs(v)} disp={wv[cell.coordinate].value!r}")

print("\n=== Portada formulas referencing Historial (first 50 rows) ===")
ws, wv = wb["Portada"], wb_v["Portada"]
for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=30):
    for cell in row:
        v = cell.value
        if isinstance(v,str) and v.startswith("=") and "historial" in v.lower():
            print(f"{cell.coordinate}: {v}")
            print(f"  disp={wv[cell.coordinate].value!r}")

# merged cells with Marca
for sn in ["Calculos", "Portada"]:
    ws = wb[sn]
    print(f"\n=== {sn} merged ranges touching Marca labels ===")
    for mc in ws.merged_cells.ranges:
        top = ws.cell(mc.min_row, mc.min_col)
        if top.value and isinstance(top.value,str) and "marca" in top.value.lower():
            print(f"  merge {mc} top={top.coordinate} val={top.value!r}")

# named ranges marca
print("\nDefined names containing marca/equipo/cert:")
for dn, defn in wb.defined_names.items():
    if any(k in dn.lower() for k in ["marca","equipo","cert","cliente","historial"]):
        print(f"  {dn}: {defn.attr_text}")
