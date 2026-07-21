import openpyxl
from pathlib import Path
path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
ws = wb["Calculos"]
for ref in ["J9","J10","J11","J12","F10","A15","B15","C15","D15","E15","F15","G15","H15","I15","J15","K15","AE28"]:
    print(f"{ref}: {ws[ref].value!r}")
print("--- Patrones E-I rows 4-6 ---")
p = wb["Patrones"]
for r in range(4, 7):
    for col in "EFGHI":
        print(f"{col}{r}: {p[f'{col}{r}'].value!r}")
print("--- VLOOKUP Patrones col 4 ---")
for sn in wb.sheetnames:
    ws2 = wb[sn]
    for row in ws2.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "VLOOKUP" in v.upper() and "Patrones" in v:
                import re
                if re.search(r",\s*4\s*,", v):
                    print(f"{sn}!{cell.coordinate}: {v}")
            if isinstance(v, str) and "Patrones!" in v and "$I$" in v:
                print(f"I-col {sn}!{cell.coordinate}: {v}")
print("--- Q column EMP source sample Q28:Q32 ---")
for r in range(28, 33):
    print(f"Q{r}: {ws[f'Q{r}'].value!r}")
wb.close()
