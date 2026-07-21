import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from openpyxl.utils import get_column_letter
path = r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm"
wb = openpyxl.load_workbook(path, data_only=False)
wbv = openpyxl.load_workbook(path, data_only=True)
ws, wv = wb["Calculos"], wbv["Calculos"]
print("Calculos rows 1-7 (cert + cliente):")
for r in range(1,8):
    parts=[]
    for c in range(1,16):
        v=ws.cell(r,c).value
        if v is None: continue
        a=get_column_letter(c)+str(r)
        parts.append(f"{a}={v!r}")
    if parts: print("R"+str(r), " | ".join(parts))
print("\nCalculos row 3 cols N-S:")
for c in range(14,20):
    a=get_column_letter(c)+"3"
    print(a, ws[a].value, "disp", wbv["Calculos"][a].value)
print("\nPortada instrument block ~row 18-28:")
ws2, wv2 = wb["Portada"], wbv["Portada"]
for r in range(18,29):
    parts=[]
    for c in range(1,12):
        v=ws2.cell(r,c).value
        if v is None: continue
        a=get_column_letter(c)+str(r)
        k="F" if isinstance(v,str) and v.startswith("=") else "C"
        parts.append(f"{a}({k})={v!r}")
    if parts: print("R"+str(r), " | ".join(parts))
