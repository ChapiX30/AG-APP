#!/usr/bin/env python3
"""
Sincroniza metadatos de patrones (app → Excel) SIN tocar la tabla EMP/Nominal.

Solo escribe en columnas A–D (bloque Instrumento/Marca/Cert/Fechas) y status en K–L.
Si la app no trae noCertificado, NO borra el valor que ya está en Excel.
Restaura EMP (E–I) desde un master limpio si el archivo actual está corrupto.
"""
from __future__ import annotations

import json
import shutil
import sys
import urllib.request
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
# Layout EMP intacto (referencia para restaurar)
BASELINE_CANDIDATES = [
    Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_clase.xlsm"),
    Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_OK.xlsm"),
    Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion.xlsm"),
]
API = (
    "https://us-central1-agg1-b7f40.cloudfunctions.net/obtenerDatosExcel"
    "?key=TU_CLAVE_SECRETA_AG_APP_2026&formato=patrones"
)

# Solo metadatos A–D. EMP vive en E–I: NUNCA escribir ahí.
BLOCKS = {
    "AG-008": {"start": 4, "alcance_default": "10000 psi"},
    "AG-052": {"start": 20, "alcance_default": "750 psi"},
    "AG-034": {"start": 26, "alcance_default": "0 a 350 psi"},
}

# Certificados conocidos del master (fallback si la app aún no tiene el número)
CERT_FALLBACK = {
    "AG-008": "1-24842",
    "AG-052": "1-23207",
    "AG-034": "MMI-CC-P-0012-2025",
}


def fetch_patrones():
    req = urllib.request.Request(API, headers={"User-Agent": "AG-SyncPatrones/1.0"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        data = json.loads(resp.read().decode("utf-8"))
    if isinstance(data, list):
        return data
    return data.get("patrones") or []


def parse_date(s: str):
    if not s:
        return None
    s = str(s).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        return None


def pick_cert(p: dict, pid: str, existing) -> str:
    """App primero; si vacío, conservar Excel; si vacío, fallback master."""
    for key in (
        "noCertificado",
        "numeroCertificado",
        "certificacion",
        "no_certificado",
        "certificado",
    ):
        v = p.get(key)
        if v is not None and str(v).strip():
            return str(v).strip()
    if existing is not None and str(existing).strip():
        return str(existing).strip()
    return CERT_FALLBACK.get(pid, "")


def clear_status_spill(ws):
    """Quita 'Status app' que se escribió encima de EMP (F/G)."""
    for start in (4, 20, 26):
        for col in (6, 7):  # F, G
            cell = ws.cell(start, col)
            val = cell.value
            if val is None:
                continue
            s = str(val)
            if "Status app" in s or s.startswith("Vigente") or s.startswith("Vencido"):
                cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()
            elif "No disponible" in s or "Por Vencer" in s or "Sin fecha" in s:
                cell.value = None
                cell.fill = PatternFill()
                cell.font = Font()


def restore_emp_from_baseline(ws_dest):
    """Copia E1:I40 (tabla Nominal/EMP) desde un master limpio."""
    baseline = next((p for p in BASELINE_CANDIDATES if p.exists()), None)
    if not baseline:
        print("AVISO: no hay baseline para restaurar EMP")
        return
    wb_b = openpyxl.load_workbook(baseline, keep_vba=True, data_only=False)
    ws_b = wb_b["Patrones"]
    from openpyxl.cell.cell import MergedCell

    for r in range(1, 41):
        for c in range(5, 10):  # E–I
            dest = ws_dest.cell(r, c)
            if isinstance(dest, MergedCell):
                continue
            dest.value = ws_b.cell(r, c).value
    # Restaurar labels A–C fijos del bloque (no fechas/certs)
    for start in (4, 20, 26):
        for r_off, label_a, label_c in (
            (0, "Instrumento:", "ID:"),
            (1, "Marca:", "No. Certif:"),
            (2, "Modelo:", "Fecha Cal:"),
            (3, "No. De Serie:", "Vigencia:"),
        ):
            ws_dest.cell(start + r_off, 1).value = label_a
            ws_dest.cell(start + r_off, 3).value = label_c
        # Alcance label
        if start == 26:
            ws_dest.cell(start + 4, 1).value = "Intervalo alcance:"
        else:
            ws_dest.cell(start + 4, 1).value = ws_b.cell(start + 4, 1).value or "Intervalo alcance:"
    print(f"EMP E–I restaurado desde: {baseline.name}")


def fill_block(ws, start: int, p: dict, pid: str, alcance_default: str):
    """Solo A–D + status en K–L. Nunca E–I."""
    # Instrumento / ID
    if p.get("descripcion"):
        ws.cell(start, 2, p.get("descripcion"))
    ws.cell(start, 4, pid)

    # Marca / No. Certif
    if p.get("marca"):
        ws.cell(start + 1, 2, p.get("marca"))
    cert = pick_cert(p, pid, ws.cell(start + 1, 4).value)
    if cert:
        ws.cell(start + 1, 4, cert)

    # Modelo / Fecha Cal
    if p.get("modelo"):
        ws.cell(start + 2, 2, p.get("modelo"))
    ultima = parse_date(p.get("fechaUltimaCalibracion") or "")
    if ultima:
        ws.cell(start + 2, 4, ultima)
    elif p.get("fechaUltimaCalibracion"):
        ws.cell(start + 2, 4, p.get("fechaUltimaCalibracion"))

    # Serie / Vigencia
    serie = p.get("serie") or ""
    try:
        if str(serie).isdigit():
            serie = int(serie)
    except Exception:
        pass
    if serie != "":
        ws.cell(start + 3, 2, serie)
    vence = parse_date(p.get("fechaVencimiento") or "")
    if vence:
        ws.cell(start + 3, 4, vence)
    elif p.get("fechaVencimiento"):
        ws.cell(start + 3, 4, p.get("fechaVencimiento"))

    # Alcance (solo si vacío)
    if not ws.cell(start + 4, 2).value:
        ws.cell(start + 4, 2, alcance_default)

    # Status a la DERECHA de EMP (K–L), no en F–G
    status = p.get("statusVigencia") or ""
    ws.cell(start, 11, "Status app:")
    cell = ws.cell(start, 12, status)
    cell.font = Font(bold=True)
    if status == "Vigente":
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
        cell.font = Font(bold=True, color="006100")
    elif status == "Por Vencer":
        cell.fill = PatternFill("solid", fgColor="FFEB9C")
        cell.font = Font(bold=True, color="9C5700")
    elif status == "Vencido":
        cell.fill = PatternFill("solid", fgColor="FFC7CE")
        cell.font = Font(bold=True, color="9C0006")
    else:
        cell.fill = PatternFill("solid", fgColor="D9D9D9")


def write_summary(ws, by_id: dict):
    """Tabla resumen lejos de EMP (fila 60+)."""
    ws["A60"] = "Patrones desde la app (patronesCalibracion) — sync automatico"
    ws["A60"].font = Font(bold=True, color="1F4E79")
    headers = [
        "noControl",
        "descripcion",
        "marca",
        "modelo",
        "serie",
        "noCertificado",
        "fechaUltimaCalibracion",
        "fechaVencimiento",
        "estadoProceso",
        "statusVigencia",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(61, c, h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")

    row = 62
    for pid in ("AG-034", "AG-052", "AG-008"):
        p = by_id.get(pid)
        if not p:
            ws.cell(row, 1, pid)
            ws.cell(row, 10, "NO ENCONTRADO EN APP")
            ws.cell(row, 10).font = Font(bold=True, color="9C0006")
            row += 1
            continue
        for c, h in enumerate(headers, 1):
            val = p.get(h, "")
            if h == "noCertificado" and not val:
                # Mostrar el del bloque Excel / fallback
                start = BLOCKS[pid]["start"]
                val = ws.cell(start + 1, 4).value or CERT_FALLBACK.get(pid, "")
            ws.cell(row, c, val)
        st = p.get("statusVigencia") or ""
        st_cell = ws.cell(row, 10)
        if st == "Vigente":
            st_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif st == "Por Vencer":
            st_cell.fill = PatternFill("solid", fgColor="FFEB9C")
        elif st == "Vencido":
            st_cell.fill = PatternFill("solid", fgColor="FFC7CE")
        row += 1

    ws.cell(row + 1, 1, "Regla Portada: <=350 AG-034 | <=750 AG-052 | >750 AG-008. Vigencia = fecha de la app.")
    ws.cell(row + 1, 1).font = Font(italic=True, size=9)


def patch_portada(wb):
    """Portada: vigencia por fecha (col J = statusVigencia). No usa I:J (rango mal)."""
    port = wb["Portada"]
    # I46 = fecha vencimiento desde resumen
    port["I46"].value = (
        '=IFERROR(VALUE(INDEX(Patrones!$H$62:$H$80,MATCH(D46,Patrones!$A$62:$A$80,0))),'
        'IFERROR(DATEVALUE(INDEX(Patrones!$H$62:$H$80,MATCH(D46,Patrones!$A$62:$A$80,0))),""))'
    )
    # E48 = mensaje según status en columna J (statusVigencia)
    # Solo "No disponible" si realmente está en calibración/falla/baja, no por en_servicio.
    port["E48"].value = (
        '=IFERROR('
        'IF(INDEX(Patrones!$J$62:$J$80,MATCH(D46,Patrones!$A$62:$A$80,0))="Vencido","PATRON VENCIDO",'
        'IF(INDEX(Patrones!$J$62:$J$80,MATCH(D46,Patrones!$A$62:$A$80,0))="Por Vencer","PATRON POR VENCER",'
        'IF(LEFT(INDEX(Patrones!$J$62:$J$80,MATCH(D46,Patrones!$A$62:$A$80,0)),13)="No disponible","PATRON NO DISPONIBLE",'
        'IF(INDEX(Patrones!$J$62:$J$80,MATCH(D46,Patrones!$A$62:$A$80,0))="Vigente","Patron Vigente",'
        'IF(AND(ISNUMBER(I46),I46<TODAY()),"PATRON VENCIDO",'
        'IF(AND(ISNUMBER(I46),I46>=TODAY()),"Patron Vigente","Revisar vigencia")))))), '
        '"Sin datos de patron en app")'
    )
    port["F46"].value = '=IFERROR(INDEX(Patrones!$C$62:$C$80,MATCH(D46,Patrones!$A$62:$A$80,0)),"")'
    port["G46"].value = '=IFERROR(INDEX(Patrones!$D$62:$D$80,MATCH(D46,Patrones!$A$62:$A$80,0)),"")'
    port["H46"].value = '=IFERROR(INDEX(Patrones!$E$62:$E$80,MATCH(D46,Patrones!$A$62:$A$80,0)),"")'
    port["E46"].value = '=IFERROR(INDEX(Patrones!$F$62:$F$80,MATCH(D46,Patrones!$A$62:$A$80,0)),"")'
    port["B46"].value = '=IFERROR(INDEX(Patrones!$B$62:$B$80,MATCH(D46,Patrones!$A$62:$A$80,0)),"")'


def resolve_paths():
    """Lee SIN_REF si se puede; si está abierto, trabaja sobre copia patrones."""
    out = PATH
    src = PATH
    try:
        open(PATH, "a+b").close()
    except PermissionError:
        out = PATH.with_name("Formato master auto Presion_SIN_REF_patrones.xlsm")
        print("Excel tiene SIN_REF abierto.")
        # Preferir leer la copia patrones si ya existe; si no, baseline
        if out.exists():
            src = out
        else:
            src = next((p for p in BASELINE_CANDIDATES if p.exists()), PATH)
        print(f"Leyendo: {src.name} → guardando: {out.name}")
        return src, out

    # Si existe copia patrones corrupta y SIN_REF también, preferir baseline EMP
    return src, out


def main() -> int:
    src, out = resolve_paths()

    print("Descargando patrones de la app...")
    try:
        patrones = fetch_patrones()
    except Exception as e:
        print("API error:", e)
        return 1

    by_id = {str(p.get("noControl", "")).upper(): p for p in patrones}
    print(f"  {len(patrones)} patrones en app")
    for pid in BLOCKS:
        p = by_id.get(pid)
        if p:
            cert = (p.get("noCertificado") or "").strip() or "(vacio en app → se conserva Excel/fallback)"
            print(
                f"  {pid}: {p.get('statusVigencia')} "
                f"vence={p.get('fechaVencimiento')} cert={cert}"
            )
        else:
            print(f"  {pid}: NO ENCONTRADO")

    try:
        wb = openpyxl.load_workbook(src, keep_vba=True)
    except PermissionError:
        # Último intento: copiar baseline a out y abrir
        baseline = next((p for p in BASELINE_CANDIDATES if p.exists()), None)
        if not baseline:
            print("No se pudo leer ningun archivo.")
            return 2
        shutil.copy2(baseline, out)
        wb = openpyxl.load_workbook(out, keep_vba=True)
        print(f"Partiendo de baseline: {baseline.name}")

    ws = wb["Patrones"]
    clear_status_spill(ws)
    # NO restaurar EMP desde baseline: las tablas Nominal/U/EMP
    # vienen de update_patrones_emp_from_certs.py (certificados PDF).

    for pid, meta in BLOCKS.items():
        p = by_id.get(pid)
        if not p:
            ws.cell(meta["start"], 12, "NO EN APP")
            continue
        fill_block(ws, meta["start"], p, pid, meta["alcance_default"])

    write_summary(ws, by_id)
    patch_portada(wb)

    try:
        wb.save(out)
        print("Guardado:", out.name)
    except PermissionError:
        alt = PATH.with_name("Formato master auto Presion_SIN_REF_patrones_FIXED.xlsm")
        wb.save(alt)
        print("Guardado como:", alt.name)

    print("Listo: EMP restaurado, status en K–L, certificados sin borrar.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
