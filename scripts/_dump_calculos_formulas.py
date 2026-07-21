import sys
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

path = r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm"
out_path = r"C:\Users\AG\Desktop\AGG\project\scripts\_calculos_formulas_out.txt"

wb = openpyxl.load_workbook(path, keep_vba=True, data_only=False)
ws = wb["Calculos"]

keywords = ("Historial", "INDEX", "VLOOKUP", "MATCH", "BD_Clientes")

def cell_formula(r, c):
    v = ws.cell(row=r, column=c).value
    if isinstance(v, str) and v.startswith("="):
        return v
    return None

lines = []
lines.append("=== Rows 1-20, cols A-M matching keywords ===")
count = 0
for r in range(1, 21):
    for c in range(1, 14):
        f = cell_formula(r, c)
        if f and any(k in f for k in keywords):
            addr = f"{get_column_letter(c)}{r}"
            lines.append(f"{addr}={f}")
            count += 1
lines.append(f"(matched {count} cells)")
lines.append("")
lines.append("=== Specific ranges ===")
specific = []
for r in range(5, 13):
    specific.append(("B", r))
for r in (5, 6):
    specific.append(("E", r))
specific.append(("F", 9))
for r in (4, 5, 6):
    specific.append(("I", r))
specific.append(("K", 4))
specific.append(("M", 8))
for col, r in specific:
    c = column_index_from_string(col)
    addr = f"{col}{r}"
    f = cell_formula(r, c)
    if f:
        lines.append(f"{addr}={f}")
    else:
        v = ws.cell(row=r, column=c).value
        lines.append(f"{addr}=(no formula, value={repr(v)})")
wb.close()
text = "\n".join(lines)
with open(out_path, "w", encoding="utf-8") as fp:
    fp.write(text)
print(text)
