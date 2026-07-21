import openpyxl
import zipfile
import re
from pathlib import Path

path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
wb_v = openpyxl.load_workbook(path, data_only=True, keep_vba=True)

def cell_info(ws, ws_v, addr):
    c = ws[addr]
    cv = ws_v[addr].value
    val = c.value
    if val is None:
        kind = "EMPTY"
    elif isinstance(val, str) and val.startswith("="):
        kind = "FORMULA"
    else:
        kind = "CONSTANT"
    return val, cv, kind

def refs_summary(formula):
    if not formula or not isinstance(formula, str) or not formula.startswith("="):
        return ""
    f = formula
    fu = formula.upper()
    parts = []
    for name in ["VLOOKUP", "INDEX", "MATCH", "XLOOKUP", "INDIRECT", "FILTER", "LET"]:
        if name in fu:
            parts.append(name)
    if "HISTORIAL" in fu or "Historial" in f:
        parts.append("Historial sheet")
    if "OBTENER" in fu:
        parts.append("ObtenerDatos/PQ")
    for m in re.finditer(r"'([^']+)'!", f):
        parts.append("'" + m.group(1) + "'!")
    for m in re.finditer(r"([A-Z]{1,3}\d+)", f):
        parts.append(m.group(1))
    seen, out = set(), []
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return ", ".join(out[:30])

print("FILE:", path.name)

for sn in ["Calculos", "Portada"]:
    if sn not in wb.sheetnames:
        print("Missing sheet", sn)
        continue
    ws, ws_v = wb[sn], wb_v[sn]
    print("\n" + "=" * 60)
    print("SHEET:", sn)
    marca_pairs = []
    for row in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=25):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip().lower() == "marca":
                for off in range(1, 6):
                    nc = ws.cell(cell.row, cell.column + off)
                    nv = ws_v.cell(cell.row, cell.column + off).value
                    if nc.value is not None or nv is not None:
                        marca_pairs.append((cell.coordinate, nc.coordinate))
                        break
    print("Marca labels -> value cells:", marca_pairs)
    for lab, addr in marca_pairs:
        val, cached, kind = cell_info(ws, ws_v, addr)
        print(f"  {addr} (label {lab}): {kind}")
        print(f"    stored: {val!r}")
        print(f"    cached/display: {cached!r}")
        if kind == "FORMULA":
            print(f"    refs: {refs_summary(val)}")

    print("Any cell in rows 1-30 containing 'marca' text:")
    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=25):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "marca" in v.lower():
                print(f"  {cell.coordinate}: {v!r}")

    print("Certificate / equipo key cells:")
    for addr in ["D4", "E4", "F4", "B9", "F9", "B4", "C4", "G4", "H4", "I4", "I5", "C9", "D9", "E9", "G9"]:
        val, cached, kind = cell_info(ws, ws_v, addr)
        if val is None and cached is None:
            continue
        line = f"  {addr}: {kind} | stored={val!r} | display={cached!r}"
        if kind == "FORMULA":
            line += " | refs=" + refs_summary(val)
        print(line)

# formulas on Calculos row 9 entire row
ws = wb["Calculos"]
print("\nCalculos row 9 (A9:Z9) formulas/constants:")
for c in range(1, 27):
    cell = ws.cell(9, c)
    if cell.value is None:
        continue
    from openpyxl.utils import get_column_letter
    addr = get_column_letter(c) + "9"
    val, cached, kind = cell_info(ws, wb_v["Calculos"], addr)
    print(f"  {addr}: {kind} {val!r} display={cached!r}")

print("\nVBA binary string search:")
with zipfile.ZipFile(path, "r") as z:
    data = z.read("xl/vbaProject.bin")
text = data.decode("latin-1", errors="ignore")
for kw in ["Worksheet_Change", "Marca", "Range(\"E4\")", "Range(\"B9\")", "Range(\"F9\")", "Historial"]:
    print(f"\n--- {kw} ---")
    idx, n = 0, 0
    while n < 6:
        i = text.find(kw, idx)
        if i < 0:
            break
        snip = re.sub(r"[\x00-\x1f]", " ", text[max(0, i - 100) : i + 150])
        print(snip[:250])
        idx = i + len(kw)
        n += 1
