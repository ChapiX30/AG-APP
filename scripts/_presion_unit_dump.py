import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
import re
import json

path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
out_path = Path(r"C:\Users\AG\Desktop\AGG\project\scripts\_presion_unit_analysis.txt")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
lines = []

def P(s=""):
    lines.append(s)

P("="*70)
P("SHEETS: " + ", ".join(wb.sheetnames))
P("="*70)

if "Calculos" in wb.sheetnames:
    ws = wb["Calculos"]
    P("\n=== UNIT CELL J10 and neighbors ===")
    for r in range(8, 13):
        for c in range(8, 13):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                P(f"  {get_column_letter(c)}{r}: {v}")

    P("\n=== F10 and pressure-related row 10 ===")
    for c in range(1, 35):
        v = ws.cell(row=10, column=c).value
        if v is not None:
            P(f"  {get_column_letter(c)}10: {v}")

    P("\n=== Rows 28-30: V,W,Y,Q,AE,N,O,F ===")
    key_cols = {"F":6,"N":14,"O":15,"Q":17,"V":22,"W":23,"Y":25,"AE":31}
    for r in [28, 29, 30]:
        P(f"\n--- Row {r} ---")
        for name, c in sorted(key_cols.items(), key=lambda x: x[1]):
            v = ws.cell(row=r, column=c).value
            P(f"  {name}{r}: {v}")
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                vu = v.upper()
                if any(k in vu for k in ["6.894", "6894", "145.038", "J$10", "$J$10", "CONVERT", "FACTOR", "KPA", "MPA", "PSI", "UNIDAD", "INDIRECT"]):
                    P(f"  CONV {get_column_letter(c)}{r}: {v}")

    P("\n=== Calculos rows 1-35: any J10 / unit / conversion ===")
    for r in range(1, 36):
        for c in range(1, 40):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("="):
                vu = v.upper()
                if "J10" in vu or "J$10" in vu or "6.894" in v or "145.038" in v:
                    P(f"  {get_column_letter(c)}{r}: {v}")

    P("\n=== Calculos A35-A80 dump (all non-empty) ===")
    for r in range(35, 81):
        parts = []
        for c in range(1, 20):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                parts.append(f"{get_column_letter(c)}{r}={str(v)[:200]}")
        if parts:
            P("  " + " | ".join(parts))

    P("\n=== All Calculos formulas referencing J10 ===")
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and ("J10" in v.upper() or "J$10" in v.upper()):
                P(f"  {cell.coordinate}: {v}")

    P("\n=== All Calculos cells with conversion factor literals ===")
    factor_pat = re.compile(r"6\.894|6894|0\.145|145\.038", re.I)
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and factor_pat.search(v):
                P(f"  {cell.coordinate}: {v}")

    P("\n=== Reading columns - sample rows 28-35 cols B-E, G-K ===")
    for r in range(28, 36):
        parts = []
        for c in [2,3,4,5,7,8,9,10,11]:
            v = ws.cell(row=r, column=c).value
            if v is not None:
                parts.append(f"{get_column_letter(c)}{r}={str(v)[:150]}")
        if parts:
            P("  " + " | ".join(parts))

for sname in wb.sheetnames:
    ws = wb[sname]
    hits = []
    for row in ws.iter_rows(max_row=120):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            vs = str(v)
            if any(k in vs.lower() for k in ["6.894", "145.038", "psi a kpa", "kpa a psi"]) or (isinstance(v, (int,float)) and abs(v-6.89476)<0.01):
                hits.append(f"{sname}!{cell.coordinate}: {vs[:200]}")
    if hits:
        P(f"\n=== Conversion hits in {sname} ===")
        for h in hits[:80]:
            P("  " + h)

P("\n=== Defined names (unit/conversion) ===")
for dn, defn in wb.defined_names.items():
    t = str(dn) + ": " + str(defn.attr_text)
    if any(k in t.lower() for k in ["unidad", "unit", "psi", "kpa", "convert", "factor", "mpa"]):
        P("  " + t)

wb.close()
out_path.write_text("\n".join(lines), encoding="utf-8")
print("Wrote", out_path, "lines", len(lines))

