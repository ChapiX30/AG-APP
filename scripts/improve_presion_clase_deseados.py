#!/usr/bin/env python3
"""
Mejora clase de exactitud → N° puntos → V. Deseado en master Presión.

Reglas (alineadas a K10/K11/K12):
  clase <= 0.6  → 8 puntos  (.2 .25 .4 .5 .6)
  clase <= 2.5  → 5 puntos  (1 1.6 2 2.5)
  clase >  2.5  → 3 puntos  (4 y 5)
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion.xlsm")

# N puntos según clase
F_PUNTOS = '=IF(OR(J11="",NOT(ISNUMBER(J11))),5,IF(J11<=0.6,8,IF(J11<=2.5,5,3)))'

# Clase: si L11 tiene valor manual, úsalo; si no, división/alcance*100
# (I11 está fusionado con H11 = etiqueta; no se puede usar)
F_CLASE = '=IF(L11="",IF(OR($J$9="",$F$10="",$F$10=0),"",$J$9/$F$10*100),L11)'

# Indicadores activos
F_K10 = '=IF(AND(ISNUMBER(J11),J11<=0.6),">> 8 pts ACTIVOS (clases 0.2; 0.25; 0.4; 0.5; 0.6)","8 pts: clases 0.2; 0.25; 0.4; 0.5; 0.6")'
F_K11 = '=IF(AND(ISNUMBER(J11),J11>0.6,J11<=2.5),">> 5 pts ACTIVOS (clases 1; 1.6; 2; 2.5)","5 pts: clases 1; 1.6; 2; 2.5")'
F_K12 = '=IF(AND(ISNUMBER(J11),J11>2.5),">> 3 pts ACTIVOS (clases 4 y 5)","3 pts: clases 4 y 5")'

# V. Deseado: primer punto 0, ultimo = alcance, intermedios redondeados a division minima
def f_deseado(col: str, row: int) -> str:
    # col letter for ROW(colrow) self-ref
    return (
        f'=IF(ROW({col}{row})-ROW(${col}$28)+1>$L$8,"",'
        f'IF(ROW({col}{row})=ROW(${col}$28),0,'
        f'IF(ROW({col}{row})-ROW(${col}$28)+1=$L$8,$F$10,'
        f'ROUND((ROW({col}{row})-ROW(${col}$28))*($F$10/($L$8-1))/$J$9,0)*$J$9)))'
    )


def main() -> int:
    if not PATH.exists():
        print(f"No existe: {PATH}")
        return 1

    print(f"Abriendo {PATH.name} ...")
    try:
        wb = openpyxl.load_workbook(PATH, keep_vba=True)
    except PermissionError:
        print("Cierra el Excel e intenta de nuevo.")
        return 1

    ws = wb["Calculos"]

    # --- Clase de exactitud ---
    # L11 = entrada manual opcional (vacía = auto desde división/alcance)
    ws["H11"].value = "Clase de Exactitud:"
    ws["J11"].value = F_CLASE
    ws["K9"].value = "Clase manual (opcional) en L11 →"
    ws["K9"].font = Font(italic=True, size=9, color="666666")
    if ws["L11"].value is None or (
        isinstance(ws["L11"].value, str) and str(ws["L11"].value).startswith("=")
    ):
        # dejar vacío para captura manual; no pisar un número que ya hayan puesto
        if not isinstance(ws["L11"].value, (int, float)):
            ws["L11"].value = None
    ws["L10"].value = "Clase manual:"
    ws["L10"].font = Font(italic=True, size=8, color="666666")

    # --- N puntos (L7 y L8 unificados; V.Deseado usa L8) ---
    ws["L7"].value = F_PUNTOS
    ws["L8"].value = F_PUNTOS
    ws["L9"].value = '=CONCATENATE(L8," puntos de calibración")'
    ws["L9"].font = Font(bold=True, color="1F4E79")

    # --- Notas K10-K12 dinámicas ---
    fill_active = PatternFill("solid", fgColor="C6EFCE")
    fill_idle = PatternFill("solid", fgColor="F2F2F2")

    ws["K10"].value = F_K10
    ws["K11"].value = F_K11
    ws["K12"].value = F_K12
    for coord in ("K10", "K11", "K12"):
        ws[coord].font = Font(size=9)
        ws[coord].fill = fill_idle

    # --- Fila 27: no debe ser la unidad (psi); dejar vacía ---
    for col in ("A", "B", "C", "D", "E", "F", "H", "I", "J", "K", "N", "O"):
        cell = ws[f"{col}27"]
        # Solo limpiar si era el arrastre de J10/psi
        if cell.value is not None:
            cell.value = None

    # --- V. Deseado / V. Generado filas 28-38 (hasta 8+ pts con margen) ---
    for row in range(28, 39):
        for col in ("A", "B", "C", "D", "E", "F"):
            ws[f"{col}{row}"].value = f_deseado(col, row)

    # Comentario corto en L6
    ws["K8"].value = "N° puntos según clase →"
    ws["K8"].font = Font(italic=True, size=9, color="666666")

    out = PATH
    try:
        wb.save(out)
        print(f"Guardado: {out}")
    except PermissionError:
        alt = PATH.with_name(PATH.stem + "_clase" + PATH.suffix)
        wb.save(alt)
        print(f"Archivo abierto. Guardado como: {alt}")

    print("Listo.")
    print("  - J11 = clase (auto div/alcance, o escribe en L11 manual)")
    print("  - L8  = 8 / 5 / 3 puntos segun clase")
    print("  - A28+ = V. Deseado automaticos")
    print("  - K10/K11/K12 indican cual banda esta activa")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
