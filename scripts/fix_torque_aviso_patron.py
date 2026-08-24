# -*- coding: utf-8 -*-
"""Arregla avisos de patrón en Torque (sin COM protection issues)."""
from __future__ import annotations

import re
import shutil
import sys
import zipfile
from io import BytesIO
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace", line_buffering=True)

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato Torque.xlsm")
MSG = "⚠ PATRÓN VENCIDO — Actualiza vigencia o cambia el patrón"

TARGETS = {
    "Portada": ["D64"],
    "Toma Datos": ["B78"],
    "Patrones": ["A8", "A17", "A25", "A33", "A39", "A45", "A51", "D53"],
}


def strip_protection(src: Path, dst: Path) -> int:
    removed = 0
    pattern = re.compile(
        rb"<sheetProtection\b[^>]*/>|<sheetProtection\b[\s\S]*?</sheetProtection>", re.I
    )
    with zipfile.ZipFile(src, "r") as zin:
        buf = BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for info in zin.infolist():
                data = zin.read(info.filename)
                if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                    data2, n = pattern.subn(b"", data)
                    if n:
                        removed += n
                        data = data2
                zout.writestr(info, data)
        dst.write_bytes(buf.getvalue())
    return removed


def fix_val(v: str | None) -> str | None:
    if not isinstance(v, str) or not v.startswith("="):
        return None
    if "vencid" not in v.lower() and "Actualiza vigencia" not in v:
        return None

    def repl(m: re.Match) -> str:
        inner = m.group(1)
        low = inner.lower()
        if "vencid" in low or "actualiza" in low or "⚠" in inner:
            m2 = re.search(r"(Vencido|VENCIDO)(\s+[\d.]+\s*Nm.*)$", inner, re.I)
            extra = m2.group(2) if m2 else ""
            return f'"{MSG}{extra}"'
        return m.group(0)

    return re.sub(r'"([^"]*)"', repl, v)


def main() -> int:
    work = PATH.with_name("_torque_aviso_work.xlsm")
    n = strip_protection(PATH, work)
    print("protections removed:", n)

    wb = openpyxl.load_workbook(work, keep_vba=True)
    fill = PatternFill("solid", fgColor="FFDCDC")
    font = Font(name="Calibri", bold=True, size=10, color="991B1B")
    align = Alignment(wrap_text=True, vertical="center")

    for sheet, addrs in TARGETS.items():
        if sheet not in wb.sheetnames:
            print("missing sheet", sheet)
            continue
        ws = wb[sheet]
        for addr in addrs:
            cell = ws[addr]
            raw = cell.value if isinstance(cell.value, str) else None
            new = fix_val(raw)
            if new and new != raw:
                cell.value = new
                cell.fill = fill
                cell.font = font
                cell.alignment = align
                print("fixed", sheet, addr)
            else:
                print("skip", sheet, addr, (raw or "")[:70])

    wb.save(work)
    wb.close()
    shutil.copy2(work, PATH)
    work.unlink(missing_ok=True)
    print("OK Torque")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
