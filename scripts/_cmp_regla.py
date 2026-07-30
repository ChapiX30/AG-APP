# -*- coding: utf-8 -*-
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

FOLDER = Path(r"C:\Users\AG\Desktop\FORMATOS AG")
MM = FOLDER / "Formato regla-flex mm.xlsx"
IN = FOLDER / "Formato regla-flex in.xlsx"

for f in (MM, IN):
    print("=" * 60, f.name, f.stat().st_size)
    wb = openpyxl.load_workbook(f, data_only=False)
    print("sheets:", wb.sheetnames)
    for s in wb.sheetnames:
        ws = wb[s]
        if hasattr(ws, "max_row"):
            print(f"  [{s}] {ws.max_row}x{ws.max_column}")
    wb.close()

mm = openpyxl.load_workbook(MM, data_only=False)
inn = openpyxl.load_workbook(IN, data_only=False)

def calc_name(wb):
    for n in wb.sheetnames:
        if "calc" in n.lower():
            return n
    return wb.sheetnames[2]

cn = calc_name(mm)
ws1, ws2 = mm[cn], inn[cn]
print("\nCALC:", cn)

# formula-only diffs
fdiffs = []
all_diffs = []
max_r = min(max(ws1.max_row or 1, ws2.max_row or 1), 150)
max_c = min(max(ws1.max_column or 1, ws2.max_column or 1), 40)
for r in range(1, max_r + 1):
    for c in range(1, max_c + 1):
        v1, v2 = ws1.cell(r, c).value, ws2.cell(r, c).value
        if v1 != v2:
            all_diffs.append((get_column_letter(c) + str(r), v1, v2))
            f1 = isinstance(v1, str) and str(v1).startswith("=")
            f2 = isinstance(v2, str) and str(v2).startswith("=")
            if f1 or f2:
                fdiffs.append((get_column_letter(c) + str(r), v1, v2))

print(f"Total diffs: {len(all_diffs)}  formula diffs: {len(fdiffs)}")
print("\n=== Formula diffs (80) ===")
for addr, v1, v2 in fdiffs[:80]:
    print(f"  {addr}: MM={repr(v1)[:110]}")
    print(f"       IN={repr(v2)[:110]}")

print("\n=== Header MM ===")
for r in range(1, 30):
    row = []
    for c in range(1, 20):
        v = ws1.cell(r, c).value
        if v is not None:
            row.append(f"{get_column_letter(c)}{r}={repr(v)[:55]}")
    if row:
        print(f"R{r}:", " | ".join(row))

print("\n=== Header IN ===")
for r in range(1, 30):
    row = []
    for c in range(1, 20):
        v = ws2.cell(r, c).value
        if v is not None:
            row.append(f"{get_column_letter(c)}{r}={repr(v)[:55]}")
    if row:
        print(f"R{r}:", " | ".join(row))

mm.close(); inn.close()
