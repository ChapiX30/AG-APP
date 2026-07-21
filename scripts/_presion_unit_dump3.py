import openpyxl
from openpyxl.utils import get_column_letter

path = r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm"
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
wp = wb["Patrones"]
ws = wb["Calculos"]

print("=== Patrones row 4 headers cols A-J ===")
for c in range(1, 11):
    print(f"  {get_column_letter(c)}4: {wp.cell(4,c).value}")

for block in [(5,16,"AG-008 default"), (32,41,"AG-034"), (44,54,"AG-052")]:
    a,b,name = block
    print(f"\n=== Patrones {name} rows {a}-{b} cols E-I ===")
    for r in range(a, b+1):
        parts = [f"{get_column_letter(c)}{r}={wp.cell(r,c).value}" for c in range(5,10)]
        print("  " + " | ".join(parts))

print("\n=== Calculos P,U,AB,AC 28-30 ===")
for r in [28,29,30]:
    for col in [16,21,28,29,30]:
        print(f"  {get_column_letter(col)}{r}: {ws.cell(r,col).value}")

print("\n=== N12 patron id ===")
print("  N12:", ws.cell(12,14).value)

print("\n=== J9 label ===")
print("  H9:", ws.cell(9,8).value, "J9:", ws.cell(9,10).value)

# Compare master without SIN_REF if exists
from pathlib import Path
alt = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion.xlsm")
if alt.exists():
    wb2 = openpyxl.load_workbook(alt, data_only=False, keep_vba=True)
    ws2 = wb2["Calculos"]
    print("\n=== Compare Presion.xlsm (non-SIN) V28/Q28 ===")
    for coord in ["V28","Y28","Q28","W28","AE28","F10","J10"]:
        col = openpyxl.utils.column_index_from_string(coord.rstrip("0123456789"))
        row = int(''.join(filter(str.isdigit, coord)))
        print(f"  {coord}: SIN={ws.cell(row,col).value} | orig={ws2.cell(row,col).value}")
    wb2.close()

wb.close()
