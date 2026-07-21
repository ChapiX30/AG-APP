import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from pathlib import Path
from openpyxl.utils import get_column_letter

path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
wb = openpyxl.load_workbook(path, data_only=False)
wbv = openpyxl.load_workbook(path, data_only=True)
ws, wv = wb["Portada"], wbv["Portada"]
print("Portada rows 40-50:")
for r in range(40, 51):
    parts = []
    for c in range(1, 15):
        cell = ws.cell(r, c)
        if cell.value is None:
            continue
        a = get_column_letter(c) + str(r)
        k = "F" if isinstance(cell.value, str) and cell.value.startswith("=") else "C"
        parts.append(f"{a}({k})={cell.value!r} disp={wv[a].value!r}")
    if parts:
        print(" R" + str(r), " | ".join(parts))

print("\nPortada formulas referencing Calculos (rows 1-80):")
for row in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=20):
    for cell in row:
        v = cell.value
        if isinstance(v, str) and v.startswith("=") and "calculos" in v.lower():
            if any(x in v.upper() for x in ["B9", "B10", "B11", "B12", "F9", "E4", "D4", "F4"]):
                print(f"  {cell.coordinate}: {v} -> {wv[cell.coordinate].value!r}")
