#!/usr/bin/env python3
"""
Patron automatico en Portada/Calculos segun alcance (F10, psi).

  alcance <= 350  → AG-034  Fluke PV350
  alcance <= 750  → AG-052  Druck DPI104
  alcance >  750  → AG-008  Omega DPG-4000-10K
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

CANDIDATES = [
    Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_clase.xlsm"),
    Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion.xlsm"),
]

# Seleccion unica (fuente de verdad en Calculos)
F_ID = '=IF(OR(F10="",NOT(ISNUMBER(F10))),"",IF(F10<=350,"AG-034",IF(F10<=750,"AG-052","AG-008")))'
F_NOM = '=IF(N12="","",IF(N12="AG-034","Fluke PV350 (0-350 psi)",IF(N12="AG-052","Druck DPI104 (0-750 psi)","Omega DPG-4000-10K (0-10000 psi)")))'

def idx(col: str, offset: int) -> str:
    """INDEX/MATCH contra ID en Patrones!D (filas 4, 20, 26)."""
    base = 'MATCH($N$12,Patrones!$D:$D,0)'
    if offset == 0:
        return f"=IF($N$12=\"\",\"\",INDEX(Patrones!{col}:{col},{base}))"
    return f"=IF($N$12=\"\",\"\",INDEX(Patrones!{col}:{col},{base}+{offset}))"


def pick_file() -> Path:
    for p in CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError("No se encontro el master Presion")


def main() -> int:
    path = pick_file()
    print(f"Archivo: {path}")
    try:
        wb = openpyxl.load_workbook(path, keep_vba=True)
    except PermissionError:
        print("Cierra el Excel e intenta de nuevo.")
        return 1

    calc = wb["Calculos"]
    port = wb["Portada"]
    pat = wb["Patrones"]

    # --- Calculos: ID seleccionado ---
    calc["L12"].value = "Patron segun alcance:"
    calc["L12"].font = Font(bold=True, size=9, color="1F4E79")
    calc["N12"].value = F_ID
    calc["N12"].font = Font(bold=True, color="006600")
    calc["M12"].value = F_NOM
    calc["M12"].font = Font(italic=True, size=9)

    # Bloque patron fila 4 = patron activo (antes fijo AG-052)
    calc["M4"].value = idx("B", 0)   # Instrumento
    calc["N4"].value = '=IF($N$12="","",$N$12)'  # ID
    calc["O4"].value = idx("D", 1)   # Certificado
    calc["P4"].value = idx("B", 1)   # Marca
    calc["Q4"].value = idx("B", 2)   # Modelo
    calc["R4"].value = idx("B", 3)   # Serie
    calc["S4"].value = idx("D", 3)   # Vigencia

    # Fila 5: ya no hardcode AG-008; vacia (un solo patron en Portada)
    for col in "MNOPQRS":
        calc[f"{col}5"].value = None

    # --- Portada: una sola fuente ---
    port["D46"].value = "=Calculos!N12"
    port["B46"].value = '=IF(D46="","",INDEX(Patrones!B:B,MATCH(D46,Patrones!D:D,0)))'
    # Array formulas → formulas normales (mas estables)
    port["E46"].value = '=IF(D46="","",INDEX(Patrones!D:D,MATCH(D46,Patrones!D:D,0)+1))'
    port["F46"].value = '=IF(D46="","",INDEX(Patrones!B:B,MATCH(D46,Patrones!D:D,0)+1))'
    port["G46"].value = '=IF(D46="","",INDEX(Patrones!B:B,MATCH(D46,Patrones!D:D,0)+2))'
    port["H46"].value = '=IF(D46="","",INDEX(Patrones!B:B,MATCH(D46,Patrones!D:D,0)+3))'
    port["I46"].value = '=IF(D46="","",INDEX(Patrones!D:D,MATCH(D46,Patrones!D:D,0)+3))'

    # Segunda fila de patron (antes logica rara >F16): limpiar
    for col in ("B", "D", "E", "F", "G", "H", "I"):
        port[f"{col}47"].value = None

    # Semaforo vigencia (ya existia)
    port["E48"].value = '=IF(OR(I46="",NOT(ISNUMBER(I46))),"",IF(I46<TODAY(),"PATRON VENCIDO","Patron Vigente"))'

    # --- Patrones: tabla de seleccion visible ---
    pat["A36"].value = "Seleccion automatica por alcance del IUT (Calculos!F10 en psi)"
    pat["A36"].font = Font(bold=True, color="1F4E79")
    pat["A37"].value = "Si alcance <="
    pat["B37"].value = "Usar ID"
    pat["C37"].value = "Equipo"
    pat["D37"].value = "Alcance en hoja"
    pat["E37"].value = "Nota fabricante"
    for c in "ABCDE":
        pat[f"{c}37"].font = Font(bold=True)
        pat[f"{c}37"].fill = PatternFill("solid", fgColor="D9E2F3")

    rows = [
        (38, 350, "AG-034", "Fluke PV350", "0 a 350 psi", "Fluke: hasta 500 psi max; ustedes usan 350"),
        (39, 750, "AG-052", "Druck DPI104", "750 psi", "Catalogo Druck: 300/1000 psi tipicos; rango lab 750"),
        (40, 10000, "AG-008", "Omega DPG-4000-10K", "10000 psi", "Omega: 0 a 10,000 psig OK"),
    ]
    for r, alc, id_, eq, alc_h, note in rows:
        pat[f"A{r}"].value = alc
        pat[f"B{r}"].value = id_
        pat[f"C{r}"].value = eq
        pat[f"D{r}"].value = alc_h
        pat[f"E{r}"].value = note

    pat["A41"].value = "Regla: <=350 → AG-034; <=750 → AG-052; >750 → AG-008"
    pat["A41"].font = Font(italic=True, size=9)

    out = path
    try:
        wb.save(out)
        print(f"Guardado: {out}")
    except PermissionError:
        alt = path.with_name(path.stem + "_patron" + path.suffix)
        wb.save(alt)
        print(f"Archivo abierto. Guardado como: {alt}")

    print("Listo: Portada D46 = Calculos!N12 segun alcance F10")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
