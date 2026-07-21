#!/usr/bin/env python3
"""
Repara master Presion:
- Error al abrir (Workbook_Open RefreshAll) → silencioso
- Fecha recepcion I4 (INDEX/MATCH por nombre de columna)
- Correo/Tel #N/D → IFERROR
IMPORTANTE: cierra Excel antes de correr.
"""
from __future__ import annotations

import json
import sys
import time
import urllib.request
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    import pythoncom
    import win32com.client
except ImportError:
    print("pip install pywin32")
    sys.exit(1)

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion.xlsm")
API = (
    "https://us-central1-agg1-b7f40.cloudfunctions.net/obtenerDatosExcel"
    "?key=TU_CLAVE_SECRETA_AG_APP_2026&prefijo=AGP"
)

HIST_HEADERS = [
    "Name", "certificado", "cliente", "equipo", "marca", "modelo", "serie",
    "id", "fecha", "tecnico", "lugarCalibracion", "frecuenciaCalibracion", "fechaRecepcion",
]

CERT = '$D$4 & "-" & TEXT($E$4,"0000") & "-" & $F$4'

F_LUGAR = f'=IFERROR(INDEX(AG_Historial[lugarCalibracion],MATCH({CERT},AG_Historial[certificado],0)),"")'

F_RECEPCION = (
    f'=IFERROR('
    f'IF(INDEX(AG_Historial[fechaRecepcion],MATCH({CERT},AG_Historial[certificado],0))="",'
    f'IF(OR(K4="Sitio",K4="sitio"),"Servicio en Sitio",""),'
    f'IFERROR(VALUE(INDEX(AG_Historial[fechaRecepcion],MATCH({CERT},AG_Historial[certificado],0))),'
    f'INDEX(AG_Historial[fechaRecepcion],MATCH({CERT},AG_Historial[certificado],0)))),'
    f'IF(OR(K4="Sitio",K4="sitio"),"Servicio en Sitio",""))'
)

F_C14 = (
    '=IF(OR(K4="Laboratorio",K4="laboratorio"),"Instalaciones AG",'
    'IF(OR(K4="Sitio",K4="sitio"),"Instalaciones de Cliente",""))'
)

VBA_OPEN = r'''
Private Sub Workbook_Open()
    On Error Resume Next
    Application.EnableEvents = True
    Application.ScreenUpdating = False
    ' RefreshAll puede fallar si Power Query pide privacidad / schema
    ThisWorkbook.RefreshAll
    DoEvents
    Application.ScreenUpdating = True
    If Err.Number <> 0 Then
        Err.Clear
        ' No MsgBox: evita "Error definido por la aplicacion o el objeto"
    End If
    On Error GoTo 0
End Sub
'''


def fetch_historial():
    req = urllib.request.Request(API, headers={"User-Agent": "AG-Repair/1.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return json.loads(resp.read().decode("utf-8")).get("historial") or []


def write_hist(ws, rows):
    for c, h in enumerate(HIST_HEADERS, 1):
        ws.cell(1, c, h)
    if ws.max_row and ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)
    for r, row in enumerate(rows, 2):
        for c, key in enumerate(HIST_HEADERS, 1):
            ws.cell(r, c, row.get(key, ""))
    n = max(len(rows) + 1, 2)
    last = chr(ord("A") + len(HIST_HEADERS) - 1)
    ref = f"A1:{last}{n}"
    if "AG_Historial" in ws.tables:
        ws.tables["AG_Historial"].ref = ref
    else:
        t = Table(displayName="AG_Historial", ref=ref)
        t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(t)
    return ref


def patch_calculos(calc):
    calc["H4"].value = "Fecha de Recepción:"
    calc["I4"].value = F_RECEPCION
    calc["K4"].value = F_LUGAR
    calc["L4"].value = "<- Lugar"
    calc["L4"].font = Font(italic=True, size=8, color="888888")
    calc["C14"].value = F_C14

    # Cliente lookups sin #N/D feo
    calc["E5"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,4,FALSE),"")'  # correo
    calc["B6"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,2,FALSE),"")'  # domicilio
    calc["E6"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,5,FALSE),"")'  # tel
    calc["B7"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,3,FALSE),"")'  # contacto


def patch_vba(path: Path) -> None:
    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0, ReadOnly=False)
    try:
        # Desactivar refresh automatico de conexiones rotas
        for i in range(1, wb.Connections.Count + 1):
            try:
                c = wb.Connections.Item(i)
                c.RefreshWithRefreshAll = False
                try:
                    c.OLEDBConnection.RefreshOnFileOpen = False
                    c.OLEDBConnection.BackgroundQuery = False
                except Exception:
                    pass
            except Exception:
                pass

        try:
            comp = wb.VBProject.VBComponents.Item("ThisWorkbook")
            code = comp.CodeModule
            try:
                start = code.ProcStartLine("Workbook_Open", 0)
                count = code.ProcCountLines("Workbook_Open", 0)
                code.DeleteLines(start, count)
            except Exception:
                pass
            code.AddFromString(VBA_OPEN)
            print("  VBA Workbook_Open seguro OK")
        except Exception as e:
            print(f"  VBA no editable: {e}")

        wb.Save()
    finally:
        wb.Close(SaveChanges=True)
        excel.Quit()
        pythoncom.CoUninitialize()


def main() -> int:
    if not PATH.exists():
        print("No existe", PATH)
        return 1

    # Probar si esta abierto
    try:
        test = open(PATH, "a+b")
        test.close()
    except PermissionError:
        print("CIERRA Excel por completo (Formato master auto Presion.xlsm) y vuelve a correr:")
        print("  python scripts\\repair_presion_open_error.py")
        return 2

    print("1) Historial + formulas (openpyxl)...")
    historial = fetch_historial()
    print(f"   filas {len(historial)}")
    wb = openpyxl.load_workbook(PATH, keep_vba=True)
    hist = None
    for n in wb.sheetnames:
        if "obtener" in n.lower():
            hist = wb[n]
            break
    if hist is None:
        hist = wb.create_sheet("obtenerDatosExcel")
    write_hist(hist, historial)
    patch_calculos(wb["Calculos"])
    wb.save(PATH)
    print("   I4=", str(wb["Calculos"]["I4"].value)[:60])
    wb.close()

    print("2) VBA / conexiones (Excel COM)...")
    time.sleep(1)
    patch_vba(PATH)
    print("LISTO. Abre el archivo de nuevo.")
    print("Si Power Query falla, los datos ya estan cargados en la hoja; no necesit as refrescar ahora.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
