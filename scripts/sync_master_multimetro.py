#!/usr/bin/env python3
"""
Sincroniza Formato Multimetro.xlsx con Firestore vía obtenerDatosExcel (prefijo AGEL).

Solo jala datos de la app (historial + clientes) y enlaza Calculos/Portada.
NO modifica bloques de incertidumbre / CMC / lecturas.

Uso:
  python scripts/sync_master_multimetro.py
  python scripts/sync_master_multimetro.py --prefijo AGEL
  python scripts/sync_master_multimetro.py --solo-sync
  python scripts/sync_master_multimetro.py --archivo "C:\\Users\\AG\\Desktop\\FORMATOS AG\\Formato Multimetro.xlsx"

Después: abre el Excel → Calculos → D4=AGEL / E4=número / F4=año → datos de hoja se llenan solos.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
import urllib.request
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Falta openpyxl. Instala con: pip install openpyxl")
    sys.exit(1)

API_URL = (
    "https://us-central1-agg1-b7f40.cloudfunctions.net/obtenerDatosExcel"
    "?key=TU_CLAVE_SECRETA_AG_APP_2026"
)
DEFAULT_MASTER = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato Multimetro.xlsx")
HIST_SHEET_NAME = "obtenerDatosExcel"

HIST_HEADERS = [
    "Name",
    "certificado",
    "cliente",
    "equipo",
    "marca",
    "modelo",
    "serie",
    "id",
    "fecha",
    "tecnico",
    "lugarCalibracion",
    "frecuenciaCalibracion",
    "fechaRecepcion",
    # Contacto del cliente (join con catálogo; evita VLOOKUP frágil por nombre)
    "domicilio",
    "contacto",
    "correo",
    "telefono",
]

CLIENTE_HEADERS = ["Nombre", "Domicilio", "Contacto", "Correo", "Telefono"]

# Catálogo de patrones (app). Solo se usan noCertificado + fechaVencimiento en hoja Patron.
PATRON_HEADERS = [
    "noControl",
    "descripcion",
    "marca",
    "modelo",
    "serie",
    "noCertificado",
    "fechaUltimaCalibracion",
    "fechaVencimiento",
    "estadoProceso",
    "statusVigencia",
    "laboratorio",
]
PS = "BD_Patrones"
# Lookup por ID en Patron!D7 → certificado (F) y vencimiento (H)
F_PATRON_CERT = (
    f'=IFERROR(INDEX({PS}!$F:$F,MATCH(TRIM($D$7),{PS}!$A:$A,0)),"")'
)
F_PATRON_VENCE = (
    f'=IFERROR(INDEX({PS}!$H:$H,MATCH(TRIM($D$7),{PS}!$A:$A,0)),"")'
)

# Clave de certificado = D4 & "-" & TEXT(E4,"0000") & "-" & F4  (igual que Presión)
# IMPORTANTE: no usar tablas estructuradas AG_Historial[...] — openpyxl/Excel las
# convierten en #REF! al reabrir. Rangos de hoja son estables.
CERT_KEY = '$D$4 & "-" & TEXT($E$4,"0000") & "-" & $F$4'
CERT_KEY_REL = 'D4 & "-" & TEXT(E4,"0000") & "-" & F4'
HS = "obtenerDatosExcel"  # hoja del historial
# Columnas: B certificado, C cliente, D equipo, E marca, F modelo, G serie,
# H id, I fecha, J tecnico, K lugar, L frecuencia, M fechaRecepcion,
# N domicilio, O contacto, P correo, Q telefono
CERT_MATCH = f"MATCH({CERT_KEY},{HS}!$B:$B,0)"
CERT_MATCH_REL = f"MATCH({CERT_KEY_REL},{HS}!$B:$B,0)"


def _idx(col: str, match: str = CERT_MATCH) -> str:
    return f"=IFERROR(INDEX({HS}!${col}:${col},{match}),\"\")"


def _idx_fallback(col: str, fallback: str, match: str = CERT_MATCH_REL) -> str:
    return f'=IFERROR(INDEX({HS}!${col}:${col},{match}),"{fallback}")'


F_CLIENTE = _idx("C")
F_EQUIPO = _idx_fallback("D", "No encontrado")
F_MARCA = _idx_fallback("E", "No encontrado")
F_MODELO = _idx("F")
F_SERIE = _idx("G")
F_ID = _idx("H")
F_FECHA_CAL = f'=IFERROR(VALUE(INDEX({HS}!$I:$I,{CERT_MATCH})),"")'
F_TECNICO = _idx("J")
F_DOMICILIO = _idx("N")
F_CONTACTO = _idx("O")
F_CORREO = _idx("P")
F_TEL = _idx("Q")
F_LUGAR = _idx("K")
F_RECEPCION = (
    "=IFERROR("
    f'IF(INDEX({HS}!$M:$M,{CERT_MATCH})="",'
    'IF(OR(UPPER(LEFT(K4,1))="S"),"Servicio en Sitio",""),'
    f"VALUE(INDEX({HS}!$M:$M,{CERT_MATCH}))),"
    'IF(OR(UPPER(LEFT(K4,1))="S"),"Servicio en Sitio",""))'
)
F_SUGERIDA = (
    f'=IFERROR(EDATE(I5, IF(INDEX({HS}!$L:$L,{CERT_MATCH})="6 meses", 6, 12)), "")'
)
F_C14 = (
    '=IF(OR(K4="Laboratorio",K4="laboratorio"),"Instalaciones AG",'
    'IF(OR(K4="Sitio",K4="sitio"),"Instalaciones de Cliente",""))'
)
F_CERT_PORTADA = '=Calculos!D4&"-"&TEXT(Calculos!E4,"0000")&"-"&Calculos!F4'


def fetch_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "AG-MultimetroSync/1.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return json.loads(resp.read().decode("utf-8"))


def col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def norm_cliente(name: str) -> str:
    """Normaliza nombre de cliente para cruzar historial ↔ catálogo."""
    s = str(name or "").strip().upper()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )
    s = re.sub(r"\([^)]*\)", " ", s)
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(
        r"\b(S A DE C V|SA DE CV|S DE RL DE CV|S DE R L DE C V|SAPI DE CV|SA|CV)\b",
        " ",
        s,
    )
    return re.sub(r"\s+", " ", s).strip()


def cliente_fields(c: dict) -> dict[str, str]:
    return {
        "Nombre": str(c.get("Nombre") or c.get("nombre") or "").strip(),
        "Domicilio": str(c.get("Domicilio") or c.get("direccion") or "").strip(),
        "Contacto": str(c.get("Contacto") or c.get("contacto") or "").strip(),
        "Correo": str(c.get("Correo") or c.get("email") or "").strip(),
        "Telefono": str(c.get("Telefono") or c.get("telefono") or "").strip(),
    }


def build_cliente_index(clientes: list[dict]) -> tuple[dict[str, dict], dict[str, dict]]:
    by_exact: dict[str, dict] = {}
    by_norm: dict[str, dict] = {}
    for raw in clientes:
        c = cliente_fields(raw)
        if not c["Nombre"]:
            continue
        by_exact.setdefault(c["Nombre"].upper(), c)
        by_norm.setdefault(norm_cliente(c["Nombre"]), c)
    return by_exact, by_norm


def lookup_cliente(
    nombre: str,
    by_exact: dict[str, dict],
    by_norm: dict[str, dict],
) -> dict | None:
    key = str(nombre or "").strip().upper()
    if not key:
        return None
    if key in by_exact:
        return by_exact[key]
    nk = norm_cliente(nombre)
    if nk and nk in by_norm:
        return by_norm[nk]
    # Prefijo: "PANASONIC…SA DE CV" ↔ "PANASONIC…SA DE CV (YENSY)"
    if nk:
        for cand_k, cand in by_norm.items():
            if cand_k.startswith(nk) or nk.startswith(cand_k):
                return cand
    return None


def enrich_historial(historial: list[dict], clientes: list[dict]) -> tuple[list[dict], int]:
    """Pega domicilio/contacto/correo/tel del catálogo en cada fila del historial."""
    by_exact, by_norm = build_cliente_index(clientes)
    matched = 0
    out: list[dict] = []
    for row in historial:
        r = dict(row)
        # Si la API ya trae columnas, respetarlas; si no, cruzar por nombre
        already = any(str(r.get(k) or "").strip() for k in ("domicilio", "contacto", "correo", "telefono"))
        if not already:
            hit = lookup_cliente(str(r.get("cliente") or ""), by_exact, by_norm)
            if hit:
                r["domicilio"] = hit["Domicilio"]
                r["contacto"] = hit["Contacto"]
                r["correo"] = hit["Correo"]
                r["telefono"] = hit["Telefono"]
                matched += 1
            else:
                r.setdefault("domicilio", "")
                r.setdefault("contacto", "")
                r.setdefault("correo", "")
                r.setdefault("telefono", "")
        else:
            matched += 1
        # Quitar espacios basura en nombre (rompe VLOOKUP legacy)
        if r.get("cliente"):
            r["cliente"] = str(r["cliente"]).strip()
        out.append(r)
    return out, matched


def clear_sheet_data(ws, start_row: int = 2) -> None:
    if ws.max_row and ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_historial(ws, rows: list[dict]) -> str:
    for col, h in enumerate(HIST_HEADERS, 1):
        ws.cell(1, col, h)
    clear_sheet_data(ws, 2)
    for r_i, row in enumerate(rows, 2):
        for c_i, key in enumerate(HIST_HEADERS, 1):
            ws.cell(r_i, c_i, row.get(key, ""))
    n = max(len(rows) + 1, 2)
    last_col = col_letter(len(HIST_HEADERS))
    return f"A1:{last_col}{n}"


def write_clientes(ws, clientes: list[dict]) -> None:
    """Escribe catálogo con nombres TRIM + filas alias del historial cuando el nombre no coincide 1:1."""
    for col, h in enumerate(CLIENTE_HEADERS, 1):
        ws.cell(1, col, h)
    clear_sheet_data(ws, 2)

    seen: set[str] = set()
    r_i = 2
    for raw in clientes:
        c = cliente_fields(raw)
        if not c["Nombre"]:
            continue
        key = c["Nombre"].upper()
        if key in seen:
            continue
        seen.add(key)
        ws.cell(r_i, 1, c["Nombre"])
        ws.cell(r_i, 2, c["Domicilio"])
        ws.cell(r_i, 3, c["Contacto"])
        ws.cell(r_i, 4, c["Correo"])
        ws.cell(r_i, 5, c["Telefono"])
        r_i += 1


def add_cliente_aliases(ws, historial: list[dict], clientes: list[dict]) -> int:
    """Agrega filas en BD_Clientes con el nombre exacto del historial si no existe."""
    by_exact, by_norm = build_cliente_index(clientes)
    existing = {
        str(ws.cell(r, 1).value or "").strip().upper()
        for r in range(2, (ws.max_row or 1) + 1)
        if ws.cell(r, 1).value
    }
    added = 0
    r_i = (ws.max_row or 1) + 1
    for row in historial:
        nombre = str(row.get("cliente") or "").strip()
        if not nombre or nombre.upper() in existing:
            continue
        hit = lookup_cliente(nombre, by_exact, by_norm)
        if not hit:
            continue
        ws.cell(r_i, 1, nombre)
        ws.cell(r_i, 2, hit["Domicilio"] or row.get("domicilio") or "")
        ws.cell(r_i, 3, hit["Contacto"] or row.get("contacto") or "")
        ws.cell(r_i, 4, hit["Correo"] or row.get("correo") or "")
        ws.cell(r_i, 5, hit["Telefono"] or row.get("telefono") or "")
        existing.add(nombre.upper())
        r_i += 1
        added += 1
    return added


def parse_fecha_patron(raw) -> object:
    """Convierte yyyy-mm-dd a datetime para que Excel lo trate como fecha."""
    from datetime import datetime

    if raw is None or raw == "":
        return ""
    if hasattr(raw, "year"):
        return raw
    s = str(raw).strip()[:10]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s


def write_patrones(
    ws,
    patrones: list[dict],
    cert_fallback_by_id: dict[str, str] | None = None,
) -> None:
    fallback = cert_fallback_by_id or {}
    for col, h in enumerate(PATRON_HEADERS, 1):
        ws.cell(1, col, h)
    clear_sheet_data(ws, 2)
    for r_i, p in enumerate(patrones, 2):
        no_control = str(p.get("noControl") or "").strip().upper()
        cert = str(p.get("noCertificado") or "").strip()
        if not cert:
            cert = fallback.get(no_control, "")
        ws.cell(r_i, 1, no_control)
        ws.cell(r_i, 2, str(p.get("descripcion") or "").strip())
        ws.cell(r_i, 3, str(p.get("marca") or "").strip())
        ws.cell(r_i, 4, str(p.get("modelo") or "").strip())
        ws.cell(r_i, 5, str(p.get("serie") or "").strip())
        ws.cell(r_i, 6, cert)
        ws.cell(r_i, 7, parse_fecha_patron(p.get("fechaUltimaCalibracion")))
        ws.cell(r_i, 8, parse_fecha_patron(p.get("fechaVencimiento")))
        ws.cell(r_i, 9, str(p.get("estadoProceso") or "").strip())
        ws.cell(r_i, 10, str(p.get("statusVigencia") or "").strip())
        ws.cell(r_i, 11, str(p.get("laboratorio") or "").strip())


def wire_patron(patron_ws) -> None:
    """Solo certificado (B7) y fecha de vencimiento (B6) desde BD_Patrones vía ID en D7."""
    patron_ws["B6"].value = F_PATRON_VENCE
    patron_ws["B7"].value = F_PATRON_CERT



def resize_or_create_table(ws, table_name: str, ref: str) -> None:
    for t in list(ws.tables.values()):
        if t.name == table_name:
            t.ref = ref
            return
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def find_historial_sheet(wb):
    for name in wb.sheetnames:
        if "obtenerdatos" in name.lower() or name.lower().startswith("obtener"):
            return wb[name]
    return wb.create_sheet(HIST_SHEET_NAME)


def unmerge_if_present(ws, ref: str) -> None:
    try:
        ws.unmerge_cells(ref)
    except Exception:
        pass


def wire_calculos(calc) -> None:
    """Enlaza cabecera de Calculos a AG_Historial / BD_Clientes (sin tocar lecturas/CMC)."""
    # Liberar D4 y F4 (layout actual: A4:D4 label, E4:F4 certificado completo)
    unmerge_if_present(calc, "A4:D4")
    unmerge_if_present(calc, "E4:F4")

    calc["A4"].value = "No. de certificado:"
    # Ejemplo que sí existe en historial AGEL (Panasonic / 200-26)
    calc["D4"].value = "AGEL"
    calc["E4"].value = 200
    calc["F4"].value = 26

    calc["B5"].value = F_CLIENTE
    calc["E5"].value = F_CORREO
    calc["B6"].value = F_DOMICILIO
    calc["E6"].value = F_TEL
    calc["B7"].value = F_CONTACTO

    calc["H4"].value = "Fecha de Recepción:"
    calc["I4"].value = F_RECEPCION
    calc["K4"].value = F_LUGAR
    calc["L4"].value = "<- Lugar (Sitio/Lab)"
    calc["L4"].font = Font(italic=True, size=8, color="888888")

    calc["I5"].value = F_FECHA_CAL
    calc["I6"].value = F_SUGERIDA
    calc["I7"].value = "=TODAY()"

    calc["B9"].value = F_EQUIPO
    calc["B10"].value = F_MARCA
    calc["B11"].value = F_MODELO
    calc["B12"].value = F_SERIE
    calc["F9"].value = F_ID

    calc["C14"].value = F_C14
    # Nombre del técnico bajo etiqueta Calibró (M7)
    calc["M8"].value = F_TECNICO


def wire_portada(portada) -> None:
    # Mostrar certificado completo compuesto (antes apuntaba solo a E4)
    if portada["J9"].value is not None or portada["H9"].value:
        portada["J9"].value = F_CERT_PORTADA


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync master Multímetro desde Firebase")
    parser.add_argument("--archivo", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--prefijo", default="AGEL", help="Filtrar certificados (AGEL). Vacío = todos")
    parser.add_argument("--url", default=API_URL)
    parser.add_argument(
        "--solo-sync",
        action="store_true",
        help="Solo refrescar hojas de datos (no reescribir fórmulas de Calculos)",
    )
    parser.add_argument(
        "--solo-wire",
        action="store_true",
        help="Solo cablear fórmulas (no descargar de la API)",
    )
    args = parser.parse_args()

    if not args.archivo.exists():
        print(f"No existe el archivo: {args.archivo}")
        return 1

    historial: list[dict] = []
    clientes: list[dict] = []
    patrones: list[dict] = []

    if not args.solo_wire:
        url = args.url
        if args.prefijo:
            sep = "&" if "?" in url else "?"
            url = f"{url}{sep}prefijo={args.prefijo}"
        print(f"Descargando datos…\n  {url}")
        data = fetch_json(url)
        historial = data.get("historial") or []
        clientes = data.get("clientes") or []
        patrones = data.get("patrones") or []
        print(
            f"  historial: {len(historial)}  clientes: {len(clientes)}  patrones: {len(patrones)}"
        )
        historial, n_join = enrich_historial(historial, clientes)
        print(f"  clientes cruzados al historial: {n_join}/{len(historial)}")

    print(f"Abriendo formato…\n  {args.archivo}")
    try:
        wb = openpyxl.load_workbook(args.archivo)
    except PermissionError:
        print("El archivo está abierto en Excel. Ciérralo e intenta de nuevo.")
        return 1

    if not args.solo_wire:
        hist_ws = find_historial_sheet(wb)
        ref = write_historial(hist_ws, historial)
        try:
            resize_or_create_table(hist_ws, "AG_Historial", ref)
        except Exception as e:
            print(f"Aviso tabla AG_Historial: {e}")

        if "BD_Clientes" in wb.sheetnames:
            cli_ws = wb["BD_Clientes"]
        else:
            cli_ws = wb.create_sheet("BD_Clientes")
        write_clientes(cli_ws, clientes)
        aliases = add_cliente_aliases(cli_ws, historial, clientes)
        if aliases:
            print(f"  alias BD_Clientes agregados: {aliases}")

        # Patrones app → BD_Patrones (cert + vencimiento para hoja Patron)
        cert_fb: dict[str, str] = {}
        if "Patron" in wb.sheetnames:
            pid = str(wb["Patron"]["D7"].value or "").strip().upper()
            pcert = wb["Patron"]["B7"].value
            if pid and pcert and not (isinstance(pcert, str) and str(pcert).startswith("=")):
                cert_fb[pid] = str(pcert).strip()
        if "BD_Patrones" in wb.sheetnames:
            pat_ws = wb["BD_Patrones"]
        else:
            pat_ws = wb.create_sheet("BD_Patrones")
        write_patrones(pat_ws, patrones, cert_fb)
        print(f"  BD_Patrones: {len(patrones)} filas")

    if not args.solo_sync:
        if "Calculos" not in wb.sheetnames:
            print("No hay hoja Calculos")
            return 1
        wire_calculos(wb["Calculos"])
        if "Portada" in wb.sheetnames:
            wire_portada(wb["Portada"])
        if "Patron" in wb.sheetnames:
            wire_patron(wb["Patron"])
            print("Patron!B6/B7 enlazados a BD_Patrones (vencimiento / certificado).")
        print("Fórmulas de cabecera enlazadas a la app (incertidumbres sin tocar).")

    out = args.archivo
    try:
        wb.save(out)
        print(f"Guardado: {out}")
    except PermissionError:
        alt = out.with_name(out.stem + "_synced" + out.suffix)
        if alt.exists() and alt.resolve() == out.resolve():
            alt = out.with_name(out.stem + "_out" + out.suffix)
        # Evitar Formato_synced_synced.xlsx
        while "_synced_synced" in alt.name:
            alt = alt.with_name(alt.name.replace("_synced_synced", "_synced"))
        wb.save(alt)
        print(f"El archivo estaba abierto. Guardado como:\n  {alt}")

    print("")
    print("Listo. Abre el Excel → en Calculos pon:")
    print("  D4 = AGEL   E4 = numero (ej. 200)   F4 = anio (ej. 26)")
    print("En Patron: D7 = ID del patrón (ej. ID04555) → B6 vencimiento y B7 certificado de la app.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
