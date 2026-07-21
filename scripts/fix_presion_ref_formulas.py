#!/usr/bin/env python3
"""
Repara #REF! en formulas: deja de usar tabla AG_Historial (rota)
y usa rangos normales en hoja 'Historial'.
Cierra Excel antes de correr.
"""
from __future__ import annotations

import json
import sys
import urllib.request
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.table import TableList

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion.xlsm")
API = (
    "https://us-central1-agg1-b7f40.cloudfunctions.net/obtenerDatosExcel"
    "?key=TU_CLAVE_SECRETA_AG_APP_2026&prefijo=AGP"
)

HEADERS = [
    "Name", "certificado", "cliente", "equipo", "marca", "modelo", "serie",
    "id", "fecha", "tecnico", "lugarCalibracion", "frecuenciaCalibracion", "fechaRecepcion",
]

# Columnas en Historial:
# B certificado, C cliente, D equipo, E marca, F modelo, G serie,
# H id, I fecha, J tecnico, K lugar, L frecuencia, M fechaRecepcion

CERT = '($D$4&"-"&TEXT($E$4,"0000")&"-"&$F$4)'


def idx(col: str) -> str:
    """INDEX/MATCH contra Historial!B certificado."""
    return f'INDEX(Historial!${col}:${col},MATCH({CERT},Historial!$B:$B,0))'


def main() -> int:
    out = PATH
    locked = False
    try:
        open(PATH, "a+b").close()
    except PermissionError:
        locked = True
        out = PATH.with_name("Formato master auto Presion_SIN_REF.xlsm")
        print("Excel tiene el archivo abierto.")
        print(f"Se creara: {out.name}")
        print("(Cierra el original y abre el nuevo)")

    print("Descargando datos...")
    req = urllib.request.Request(API, headers={"User-Agent": "AG-FixRef/1.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        historial = json.loads(resp.read().decode("utf-8")).get("historial") or []
    print(f"  {len(historial)} filas")

    try:
        wb = openpyxl.load_workbook(PATH, keep_vba=True)
    except PermissionError:
        print("No se pudo leer el archivo (bloqueado). Cierralo e intenta de nuevo.")
        return 2

    # Renombrar hoja de datos a Historial
    hist = None
    for name in list(wb.sheetnames):
        if "obtener" in name.lower() or name == "Historial":
            hist = wb[name]
            break
    if hist is None:
        hist = wb.create_sheet("Historial")
    if hist.title != "Historial":
        hist.title = "Historial"

    # Quitar tablas rotas (causa #REF! en refs estructuradas)
    try:
        tables = list(hist.tables.keys())
        for tname in tables:
            del hist.tables[tname]
            print(f"  tabla eliminada: {tname}")
    except Exception as e:
        print(f"  aviso tablas: {e}")

    # Escribir datos
    for c, h in enumerate(HEADERS, 1):
        hist.cell(1, c, h)
    if hist.max_row and hist.max_row >= 2:
        hist.delete_rows(2, hist.max_row - 1)
    for r, row in enumerate(historial, 2):
        for c, key in enumerate(HEADERS, 1):
            hist.cell(r, c, row.get(key, ""))

    calc = wb["Calculos"]

    # Lugar
    calc["K4"].value = f'=IFERROR({idx("K")},"")'
    calc["L4"].value = "<- Lugar"
    calc["L4"].font = Font(italic=True, size=8, color="888888")

    # Fecha recepcion (ya no #REF!)
    calc["H4"].value = "Fecha de Recepción:"
    calc["I4"].value = (
        f'=IFERROR('
        f'IF({idx("M")}="",'
        f'IF(OR(K4="Sitio",K4="sitio"),"Servicio en Sitio",""),'
        f'IFERROR(VALUE({idx("M")}),{idx("M")})),'
        f'IF(OR(K4="Sitio",K4="sitio"),"Servicio en Sitio",""))'
    )

    calc["C14"].value = (
        '=IF(OR(K4="Laboratorio",K4="laboratorio"),"Instalaciones AG",'
        'IF(OR(K4="Sitio",K4="sitio"),"Instalaciones de Cliente",""))'
    )

    # Cliente / fechas / equipo
    calc["B5"].value = f'=IFERROR({idx("C")},"")'
    calc["I5"].value = f'=IFERROR(VALUE({idx("I")}),"")'
    calc["I6"].value = (
        f'=IFERROR(EDATE(I5,IF({idx("L")}="6 meses",6,12)),"")'
    )
    calc["M8"].value = f'=IFERROR({idx("J")},"")'
    calc["B9"].value = f'=IFERROR({idx("D")},"No encontrado")'
    calc["F9"].value = f'=IFERROR({idx("H")},"")'
    calc["B10"].value = f'=IFERROR({idx("E")},"No encontrado")'
    calc["B11"].value = f'=IFERROR({idx("F")},"")'
    calc["B12"].value = f'=IFERROR({idx("G")},"")'

    # BD clientes sin #N/D
    calc["E5"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,4,FALSE),"")'
    calc["B6"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,2,FALSE),"")'
    calc["E6"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,5,FALSE),"")'
    calc["B7"].value = '=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,3,FALSE),"")'

    try:
        wb.save(out)
        print("Guardado:", out)
    except PermissionError:
        alt = PATH.with_name("Formato master auto Presion_SIN_REF.xlsm")
        wb.save(alt)
        print("Guardado como:", alt)

    print("I4=", calc["I4"].value[:80])
    print("Listo: formulas sin tabla AG_Historial (sin #REF!)")
    if locked or out != PATH:
        print("Abre:", out.name)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
