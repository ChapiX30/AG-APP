#!/usr/bin/env python3
"""
Sincroniza Formato master auto Presion.xlsm con Firestore vía obtenerDatosExcel.

Uso:
  python scripts/sync_master_presion.py
  python scripts/sync_master_presion.py --prefijo AGP
  python scripts/sync_master_presion.py --archivo "C:\\Users\\AG\\Desktop\\FORMATOS AG\\Formato master auto Presion.xlsm"

Después: abre el Excel, pon D4=AGP / E4=número / F4=año → Calculos se llena solo.
"""

from __future__ import annotations

import argparse
import json
import sys
import urllib.request
from pathlib import Path

try:
    import openpyxl
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("Falta openpyxl. Instala con: pip install openpyxl")
    sys.exit(1)

API_URL = (
    "https://us-central1-agg1-b7f40.cloudfunctions.net/obtenerDatosExcel"
    "?key=TU_CLAVE_SECRETA_AG_APP_2026"
)
DEFAULT_MASTER = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master Presion.xlsm")

HIST_HEADERS = [
    "Name",
    "certificado",
    "cliente",
    "equipo",
    "marca",
    "modelo",
    "serie",
    "id",
    "fecha",
    "tecnico",
    "lugarCalibracion",
    "frecuenciaCalibracion",
    "fechaRecepcion",
]

CLIENTE_HEADERS = ["Nombre", "Domicilio", "Contacto", "Correo", "Telefono"]


def fetch_json(url: str) -> dict:
    req = urllib.request.Request(url, headers={"User-Agent": "AG-MasterSync/1.0"})
    with urllib.request.urlopen(req, timeout=180) as resp:
        return json.loads(resp.read().decode("utf-8"))


def clear_sheet_data(ws, start_row: int = 2) -> None:
    if ws.max_row and ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def write_historial(ws, rows: list[dict]) -> str:
    """Escribe filas y devuelve el ref de tabla (ej. A1:L100)."""
    for col, h in enumerate(HIST_HEADERS, 1):
        ws.cell(1, col, h)
    clear_sheet_data(ws, 2)
    for r_i, row in enumerate(rows, 2):
        for c_i, key in enumerate(HIST_HEADERS, 1):
            ws.cell(r_i, c_i, row.get(key, ""))
    n = max(len(rows) + 1, 2)
    last_col = chr(ord("A") + len(HIST_HEADERS) - 1)
    return f"A1:{last_col}{n}"


def write_clientes(ws, clientes: list[dict]) -> None:
    for col, h in enumerate(CLIENTE_HEADERS, 1):
        ws.cell(1, col, h)
    clear_sheet_data(ws, 2)
    for r_i, c in enumerate(clientes, 2):
        ws.cell(r_i, 1, c.get("Nombre") or c.get("nombre") or "")
        ws.cell(r_i, 2, c.get("Domicilio") or c.get("direccion") or "")
        ws.cell(r_i, 3, c.get("Contacto") or c.get("contacto") or "")
        ws.cell(r_i, 4, c.get("Correo") or c.get("email") or "")
        ws.cell(r_i, 5, c.get("Telefono") or c.get("telefono") or "")


def resize_or_create_table(wb, ws, table_name: str, ref: str) -> None:
    existing = None
    for t in list(ws.tables.values()):
        if t.name == table_name:
            existing = t
            break
    if existing is not None:
        existing.ref = ref
        return
    # Si la tabla está en otra hoja / perdida, crear en esta
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def find_historial_sheet(wb):
    for name in wb.sheetnames:
        if "obtenerdatos" in name.lower() or name.lower().startswith("obtener"):
            return wb[name]
    # fallback: crear
    return wb.create_sheet("obtenerDatosExcel")


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync master Presión desde Firebase")
    parser.add_argument("--archivo", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--prefijo", default="AGP", help="Filtrar certificados (AGP). Vacío = todos")
    parser.add_argument("--url", default=API_URL)
    args = parser.parse_args()

    if not args.archivo.exists():
        print(f"No existe el archivo: {args.archivo}")
        return 1

    url = args.url
    if args.prefijo:
        sep = "&" if "?" in url else "?"
        url = f"{url}{sep}prefijo={args.prefijo}"

    print(f"Descargando datos…\n  {url}")
    data = fetch_json(url)
    historial = data.get("historial") or []
    clientes = data.get("clientes") or []
    print(f"  historial: {len(historial)}  clientes: {len(clientes)}")

    print(f"Abriendo master (conserva macros)…\n  {args.archivo}")
    wb = openpyxl.load_workbook(args.archivo, keep_vba=True)

    hist_ws = find_historial_sheet(wb)
    ref = write_historial(hist_ws, historial)
    try:
        resize_or_create_table(wb, hist_ws, "AG_Historial", ref)
    except Exception as e:
        print(f"Aviso tabla AG_Historial: {e}")
        print("  (Las fórmulas VLOOKUP seguirán funcionando si el nombre de tabla ya existe en Excel)")

    # Las fórmulas de Calculos leen la hoja plana "Historial": mantenerla en sync también
    if "Historial" in wb.sheetnames:
        write_historial(wb["Historial"], historial)
        print(f"  hoja Historial actualizada: {len(historial)} filas")

    if "BD_Clientes" in wb.sheetnames:
        write_clientes(wb["BD_Clientes"], clientes)
    else:
        ws = wb.create_sheet("BD_Clientes")
        write_clientes(ws, clientes)

    out = args.archivo
    # Guardar con sufijo _synced si el original está abierto / bloqueado
    try:
        wb.save(out)
        print(f"Guardado: {out}")
    except PermissionError:
        alt = out.with_name(out.stem + "_synced" + out.suffix)
        wb.save(alt)
        print(f"El archivo estaba abierto. Guardado como:\n  {alt}")
        print("Cierra el master original y renombra si quieres.")

    print("")
    print("Listo. Abre el Excel -> en Calculos pon:")
    print("  D4 = AGP   E4 = numero (ej. 594)   F4 = anio (ej. 26)")
    print("Los datos de la hoja de trabajo deben aparecer solos.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
