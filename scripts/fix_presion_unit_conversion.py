#!/usr/bin/env python3
"""
Hace que incertidumbre/sesgo/selección de patrón se conviertan con la unidad (J10).

Modelo:
  - Tablas Patrones F:I siempre en PSI (certificados).
  - Usuario trabaja en la unidad de J10 (F10, J9, lecturas).
  - J13 = factor (psi → unidad).
  - VLOOKUP usa clave en psi (valor/J13) y multiplica resultado × J13.
  - VBA solo actualiza etiquetas; YA NO reescribe EMP (rompía AG-034/052).
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Font

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
ALT = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF_unidades.xlsm")

# Rangos EMP actuales (update_patrones_emp_from_certs.py)
R008 = (6, 16)
R034 = (32, 41)
R052 = (44, 54)

FACTOR_FORMULA = (
    '=IFERROR('
    'IF(LOWER(TRIM(J10))="psi",1,'
    'IF(LOWER(TRIM(J10))="kpa",6.894757,'
    'IF(LOWER(TRIM(J10))="mpa",0.00689476,'
    'IF(LOWER(TRIM(J10))="bar",0.0689476,'
    'IF(LOWER(TRIM(J10))="mbar",68.94757,'
    'IF(LOWER(TRIM(J10))="kg-cm2",0.070307,'
    'IF(LOWER(TRIM(J10))="inhg",2.036021,'
    'IF(LOWER(TRIM(J10))="inh2o",27.6799,'
    'IF(LOWER(TRIM(J10))="mmhg",51.7149,1))))))))),1)'
)

# Alcance en psi para elegir patrón (soporta F10 en cualquier unidad + vacío N13)
N12_FORMULA = (
    '=IF(OR(F10="",NOT(ISNUMBER(F10))),"",'
    'IF(OR(AND(ISNUMBER(N13),N13<0),AND(ISNUMBER(F10),F10<0)),"AG-034",'
    'IF(ABS(F10)/$J$13<=350,"AG-034",'
    'IF(ABS(F10)/$J$13<=750,"AG-052","AG-008"))))'
)


def v_formula(row: int) -> str:
    f8, l8 = R008
    f34, l34 = R034
    f52, l52 = R052
    # Clave en psi = O/J13 ; resultado U en unidad = U_psi * J13
    return (
        f'=IF(OR(O{row}="",$J$13=0),"",'
        f'IF($N$12="AG-034",'
        f'IFERROR(VLOOKUP(O{row}/$J$13,Patrones!$F${f34}:$I${l34},3,TRUE)*$J$13,""),'
        f'IF($N$12="AG-052",'
        f'IFERROR(VLOOKUP(O{row}/$J$13,Patrones!$F${f52}:$I${l52},3,TRUE)*$J$13,""),'
        f'IFERROR(VLOOKUP(O{row}/$J$13,Patrones!$F${f8}:$I${l8},3,TRUE)*$J$13,""))))'
    )


def y_formula(row: int) -> str:
    f8, l8 = R008
    f34, l34 = R034
    f52, l52 = R052
    return (
        f'=IF(OR(O{row}="",$J$13=0),"",'
        f'IF($N$12="AG-034",'
        f'IFERROR(VLOOKUP(O{row}/$J$13,Patrones!$F${f34}:$I${l34},2,TRUE)*$J$13,""),'
        f'IF($N$12="AG-052",'
        f'IFERROR(VLOOKUP(O{row}/$J$13,Patrones!$F${f52}:$I${l52},2,TRUE)*$J$13,""),'
        f'IFERROR(VLOOKUP(O{row}/$J$13,Patrones!$F${f8}:$I${l8},2,TRUE)*$J$13,""))))'
    )


VBA_CHANGE = r'''
Private Sub Worksheet_Change(ByVal Target As Range)
    On Error GoTo ErrorHandler

    ' --- Cambio de certificado: restaurar formulas (evita romper formato) ---
    If Not Intersect(Target, Range("D4:F4")) Is Nothing Then
        Application.EnableEvents = False
        Call RestablecerFormulasDesdeHistorial
        Application.EnableEvents = True
        Exit Sub
    End If

    ' --- Cambio de unidad ---
    If Intersect(Target, Range("J10")) Is Nothing Then Exit Sub

    Application.EnableEvents = False

    Dim unidad As String, prev As String
    Dim fNew As Double, fOld As Double, ratio As Double
    Dim r As Long, c As Long, v As Variant

    unidad = LCase(Trim(CStr(Range("J10").Value)))
    prev = LCase(Trim(CStr(Range("J14").Value)))
    If prev = "" Then prev = "psi"

    fNew = FactorUnidad(unidad)
    fOld = FactorUnidad(prev)
    If fOld = 0 Then fOld = 1
    ratio = fNew / fOld

    If ratio <> 1 Then
        If IsNumeric(Range("F10").Value) Then
            Range("F10").Value = CDbl(Range("F10").Value) * ratio
        End If
        If IsNumeric(Range("J9").Value) Then
            Range("J9").Value = CDbl(Range("J9").Value) * ratio
        End If
        For r = 28 To 38
            For c = 3 To 6
                If Not Cells(r, c).HasFormula Then
                    v = Cells(r, c).Value
                    If IsNumeric(v) Then
                        If CDbl(v) <> 0 Then Cells(r, c).Value = CDbl(v) * ratio
                    End If
                End If
            Next c
            For c = 8 To 11
                If Not Cells(r, c).HasFormula Then
                    v = Cells(r, c).Value
                    If IsNumeric(v) Then
                        If CDbl(v) <> 0 Then Cells(r, c).Value = CDbl(v) * ratio
                    End If
                End If
            Next c
        Next r
        If Not Range("N13").HasFormula Then
            If IsNumeric(Range("N13").Value) Then
                Range("N13").Value = CDbl(Range("N13").Value) * ratio
            End If
        End If
    End If

    Range("J14").Value = unidad

    On Error Resume Next
    Sheets("Patrones").Range("E4").Value = "Resultados en " & Range("J10").Value & " (EMP en psi; U/sesgo/lecturas convertidos)"
    Sheets("CMC").Range("I1").Value = Range("J10").Value
    On Error GoTo ErrorHandler

Salir:
    Application.EnableEvents = True
    Exit Sub

ErrorHandler:
    Application.EnableEvents = True
End Sub

Private Sub RestablecerFormulasDesdeHistorial()
    ' Siempre reescribe formulas al cambiar certificado.
    ' Asi Marca/Modelo/etc. vuelven automaticos aunque un tecnico las haya escrito a mano.
    Dim k As String
    k = "($D$4&""-""&TEXT($E$4,""0000"")&""-""&$F$4)"

    Range("B5").Formula = "=IFERROR(INDEX(Historial!$C:$C,MATCH(" & k & ",Historial!$B:$B,0)),"""")"
    Range("B6").Formula = "=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,2,FALSE),"""")"
    Range("B7").Formula = "=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,3,FALSE),"""")"
    Range("E5").Formula = "=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,4,FALSE),"""")"
    Range("E6").Formula = "=IFERROR(VLOOKUP(B5,BD_Clientes!A:H,5,FALSE),"""")"

    Range("B9").Formula = "=IFERROR(INDEX(Historial!$D:$D,MATCH(" & k & ",Historial!$B:$B,0)),""No encontrado"")"
    Range("F9").Formula = "=IFERROR(INDEX(Historial!$H:$H,MATCH(" & k & ",Historial!$B:$B,0)),"""")"
    Range("B10").Formula = "=IFERROR(INDEX(Historial!$E:$E,MATCH(" & k & ",Historial!$B:$B,0)),""No encontrado"")"
    Range("B11").Formula = "=IFERROR(INDEX(Historial!$F:$F,MATCH(" & k & ",Historial!$B:$B,0)),"""")"
    Range("B12").Formula = "=IFERROR(INDEX(Historial!$G:$G,MATCH(" & k & ",Historial!$B:$B,0)),"""")"

    Range("K4").Formula = "=IFERROR(INDEX(Historial!$K:$K,MATCH(" & k & ",Historial!$B:$B,0)),"""")"
    Range("I4").Formula = "=IFERROR(IF(INDEX(Historial!$M:$M,MATCH(" & k & ",Historial!$B:$B,0))="""",IF(OR(K4=""Sitio"",K4=""sitio""),""Servicio en Sitio"",""""),IFERROR(VALUE(INDEX(Historial!$M:$M,MATCH(" & k & ",Historial!$B:$B,0))),INDEX(Historial!$M:$M,MATCH(" & k & ",Historial!$B:$B,0)))),IF(OR(K4=""Sitio"",K4=""sitio""),""Servicio en Sitio"",""""))"
    Range("I5").Formula = "=IFERROR(VALUE(INDEX(Historial!$I:$I,MATCH(" & k & ",Historial!$B:$B,0))),"""")"
    Range("I6").Formula = "=IFERROR(EDATE(I5,IF(INDEX(Historial!$L:$L,MATCH(" & k & ",Historial!$B:$B,0))=""6 meses"",6,12)),"""")"
    Range("M8").Formula = "=IFERROR(INDEX(Historial!$J:$J,MATCH(" & k & ",Historial!$B:$B,0)),"""")"

    ' Portada: espejo del instrumento (por si alguien borro formulas alla)
    On Error Resume Next
    With Sheets("Portada")
        .Range("D20").Formula = "=Calculos!B9"
        .Range("I20").Formula = "=Calculos!F9"
        .Range("D22").Formula = "=Calculos!B10"
        .Range("D24").Formula = "=Calculos!B11"
        .Range("D26").Formula = "=Calculos!B12"
        .Range("C20").Formula = "=Calculos!A9"
        .Range("C22").Formula = "=Calculos!A10"
        .Range("C24").Formula = "=Calculos!A11"
        .Range("C26").Formula = "=Calculos!A12"
    End With
    On Error GoTo 0
End Sub

Private Function FactorUnidad(ByVal u As String) As Double
    Select Case LCase(Trim(u))
        Case "psi": FactorUnidad = 1
        Case "kpa": FactorUnidad = 6.894757
        Case "mpa": FactorUnidad = 0.00689476
        Case "bar": FactorUnidad = 0.0689476
        Case "mbar": FactorUnidad = 68.94757
        Case "kg-cm2": FactorUnidad = 0.070307
        Case "inhg": FactorUnidad = 2.036021
        Case "inh2o": FactorUnidad = 27.6799
        Case "mmhg": FactorUnidad = 51.7149
        Case Else: FactorUnidad = 1
    End Select
End Function
'''


def patch_openpyxl(path: Path, out: Path) -> None:
    wb = openpyxl.load_workbook(path, keep_vba=True)
    calc = wb["Calculos"]

    calc["H13"] = "Factor (psi→unid)"
    calc["H13"].font = Font(size=8, italic=True)
    calc["J13"] = FACTOR_FORMULA
    calc["J13"].font = Font(bold=True, size=9)
    calc["K13"] = '=IF(J13=1,"psi base",J10)'
    calc["K13"].font = Font(size=8, color="666666")

    # Unidad anterior (para VBA al cambiar J10)
    calc["H14"] = "Unidad previa (VBA)"
    calc["H14"].font = Font(size=8, italic=True, color="888888")
    if not calc["J14"].value:
        calc["J14"] = "psi"

    calc["M12"] = "Patron ID"
    calc["M12"].font = Font(bold=True, size=9)
    calc["N12"] = N12_FORMULA

    for row in range(28, 39):
        calc.cell(row, 22).value = v_formula(row)  # V = U
        calc.cell(row, 25).value = y_formula(row)  # Y = sesgo

    calc["H15"] = (
        "Cambia J10 → VBA convierte F10 y J9. "
        "U/sesgo (V/Y) = tabla PSI × J13. Lecturas en la misma unidad."
    )
    calc["H15"].font = Font(size=7, italic=True, color="666666")
    port = wb["Portada"]
    port["D46"] = "=Calculos!N12"

    try:
        wb.save(out)
    except PermissionError:
        alt = out.with_name(out.stem + "_unidades.xlsm")
        wb.save(alt)
        print("Guardado (alt):", alt)
        return
    print("Formulas guardadas:", out.name)


def patch_vba(path: Path) -> bool:
    try:
        import pythoncom
        import win32com.client
    except ImportError:
        print("AVISO: sin win32com — VBA no actualizado.")
        return False

    pythoncom.CoInitialize()
    excel = win32com.client.DispatchEx("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    excel.EnableEvents = False
    try:
        wb = excel.Workbooks.Open(str(path.resolve()), UpdateLinks=0)
    except Exception as e:
        print("VBA: no se pudo abrir (cierra Excel):", e)
        excel.Quit()
        pythoncom.CoUninitialize()
        return False

    try:
        vbproj = wb.VBProject
        target = None
        for i in range(1, vbproj.VBComponents.Count + 1):
            comp = vbproj.VBComponents.Item(i)
            if comp.Name == "Hoja3":
                target = comp
                break
        if target is None:
            print("VBA: no Hoja3")
            wb.Close(False)
            return False

        code = target.CodeModule
        # Borrar TODO el módulo para no acumular FactorUnidad / Change duplicados
        n = code.CountOfLines
        if n > 0:
            code.DeleteLines(1, n)
        code.AddFromString(VBA_CHANGE)
        wb.Save()
        print("VBA Hoja3 reemplazado limpio (1 Change + 1 FactorUnidad).")
        wb.Close(True)
        return True
    except Exception as e:
        print("VBA error:", e)
        try:
            wb.Close(False)
        except Exception:
            pass
        return False
    finally:
        excel.Quit()
        pythoncom.CoUninitialize()


def main() -> int:
    src = PATH if PATH.exists() else ALT
    out = PATH
    try:
        open(PATH, "a+b").close()
    except PermissionError:
        out = ALT
        print("SIN_REF abierto →", out.name)

    patch_openpyxl(src, out)
    # VBA sobre el archivo guardado
    patch_vba(out if out.exists() else src)
    print("Listo. Prueba: J10=kPa → V/Y deben ser ~×6.895 los de psi.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
