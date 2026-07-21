#!/usr/bin/env python3
"""
Actualiza Nominal / Sesgo / Incertidumbre U / EMP en Patrones desde los 3 PDFs.

Columnas (Calculos VLOOKUP):
  F = Nominal (psi)
  G = Sesgo (error del patrón)     → Calculos!Y
  H = Incertidumbre U (k=2)        → Calculos!V  (W=V/2)
  I = EMP± = |sesgo| + U           (referencia)

Layout (sin pisar metadatos A–D):
  AG-008 Omega   header F5  data F6:I16    cert 1-24842
  AG-034 PV350   header F31 data F32:I41   cert MMI-CC-P-0012-2025  ← el que faltaba
  AG-052 Druck   header F43 data F44:I54   cert 1-23207
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

PATH = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
ALT = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF_emp.xlsm")

# (nominal_psi, sesgo_psi, U_k2_psi)
AG008 = [  # Omega — 1-24842
    (0.0, 0.00, 0.58),
    (1000.0, 0.24, 0.86),
    (2000.0, -0.41, 0.88),
    (3000.0, 0.48, 0.85),
    (4000.0, -0.10, 0.67),
    (5000.0, 1.35, 0.92),
    (6000.0, 1.77, 0.95),
    (7000.0, 1.77, 0.98),
    (8000.0, 0.49, 0.98),
    (9000.0, 0.95, 0.84),
    (10000.0, 2.04, 0.90),
]

AG034 = [  # Fluke PV350 — MMI-CC-P-0012-2025
    (35.0, -0.08, 0.021),
    (70.0, -0.16, 0.020),
    (105.0, -0.22, 0.027),
    (140.0, -0.27, 0.035),
    (175.0, -0.30, 0.038),
    (210.0, -0.36, 0.038),
    (245.0, -0.39, 0.039),
    (280.0, -0.41, 0.040),
    (315.0, -0.41, 0.032),
    (350.0, -0.38, 0.035),
]

AG052 = [  # Druck — 1-23207
    (0.0, 0.000, 0.058),
    (100.0, 0.044, 0.082),
    (200.0, 0.093, 0.058),
    (300.0, 0.092, 0.059),
    (400.0, 0.191, 0.059),
    (500.0, 0.199, 0.060),
    (600.0, 0.198, 0.061),
    (700.0, 0.293, 0.063),
    (800.0, 0.293, 0.064),
    (900.0, 0.297, 0.072),
    (1000.0, 0.323, 0.089),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=9)
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def safe_set(ws, r, c, value):
    cell = ws.cell(r, c)
    if isinstance(cell, MergedCell):
        return False
    cell.value = value
    return True


def clear_ei(ws, r1, r2):
    # Quitar merges que bloquean E–I (ej. A45:K48 Conversiones)
    to_unmerge = []
    for m in list(ws.merged_cells.ranges):
        if m.max_row < r1 or m.min_row > r2:
            continue
        if m.max_col < 5 or m.min_col > 9:
            continue
        to_unmerge.append(str(m))
    for ref in to_unmerge:
        try:
            ws.unmerge_cells(ref)
            print(f"  Unmerge {ref}")
        except Exception:
            pass
    for r in range(r1, r2 + 1):
        for c in range(5, 10):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell.fill = PatternFill()
            cell.border = Border()


def write_table(ws, header_row: int, title: str, rows: list, id_ref: str):
    safe_set(ws, header_row, 5, title)
    ws.cell(header_row, 5).font = Font(bold=True, size=8, color="1F4E79")

    for c, h in enumerate(("Nominal", "Sesgo", "Incert. U(k=2)", "EMP±"), start=6):
        if safe_set(ws, header_row, c, h):
            cell = ws.cell(header_row, c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN

    first = header_row + 1
    for i, (nom, sesgo, u) in enumerate(rows):
        r = first + i
        emp = round(abs(float(sesgo)) + float(u), 4)
        safe_set(ws, r, 5, f'={id_ref}&"-"&TEXT(F{r},"0")')
        safe_set(ws, r, 6, float(nom))
        safe_set(ws, r, 7, float(sesgo))
        safe_set(ws, r, 8, float(u))
        safe_set(ws, r, 9, emp)
        for c in range(5, 10):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.border = THIN
            if c >= 6:
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "0.000"
    last = first + len(rows) - 1
    return first, last


def patch_calculos(wb, r008, r034, r052):
    calc = wb["Calculos"]
    calc["M12"] = "Patron ID"
    calc["M12"].font = Font(bold=True, size=9)
    calc["N12"] = (
        '=IF(OR(F10="",NOT(ISNUMBER(F10))),"",'
        'IF(F10<=350,"AG-034",IF(F10<=750,"AG-052","AG-008")))'
    )

    f8, l8 = r008
    f34, l34 = r034
    f52, l52 = r052

    for row in range(28, 39):
        # V = U (col 3)
        calc.cell(row, 22).value = (
            f'=IF(O{row}="","",IF($N$12="AG-034",'
            f'IFERROR(VLOOKUP(O{row},Patrones!$F${f34}:$I${l34},3,TRUE),""),'
            f'IF($N$12="AG-052",'
            f'IFERROR(VLOOKUP(O{row},Patrones!$F${f52}:$I${l52},3,TRUE),""),'
            f'IFERROR(VLOOKUP(O{row},Patrones!$F${f8}:$I${l8},3,TRUE),""))))'
        )
        # Y = Sesgo (col 2)
        calc.cell(row, 25).value = (
            f'=IF(O{row}="","",IF($N$12="AG-034",'
            f'IFERROR(VLOOKUP(O{row},Patrones!$F${f34}:$I${l34},2,TRUE),""),'
            f'IF($N$12="AG-052",'
            f'IFERROR(VLOOKUP(O{row},Patrones!$F${f52}:$I${l52},2,TRUE),""),'
            f'IFERROR(VLOOKUP(O{row},Patrones!$F${f8}:$I${l8},2,TRUE),""))))'
        )


def update_meta(ws):
    safe_set(ws, 5, 4, "1-24842")
    safe_set(ws, 6, 4, datetime(2025, 2, 4))
    safe_set(ws, 8, 2, "10000 psi")

    safe_set(ws, 21, 4, "1-23207")
    safe_set(ws, 22, 4, datetime(2024, 11, 28))
    safe_set(ws, 24, 2, "1000 psi")

    safe_set(ws, 27, 4, "MMI-CC-P-0012-2025")
    safe_set(ws, 28, 4, datetime(2025, 1, 7))
    safe_set(ws, 30, 2, "0 a 350 psi")


def main() -> int:
    src = PATH if PATH.exists() else ALT
    out = PATH
    try:
        open(PATH, "a+b").close()
    except PermissionError:
        out = ALT
        print(f"SIN_REF abierto → {out.name}")

    wb = openpyxl.load_workbook(src, keep_vba=True)
    ws = wb["Patrones"]

    # Limpiar EMP viejo en E5:I55 (no toca A–D ni resumen fila 60+)
    clear_ei(ws, 5, 55)
    # Restaurar "Resultados en psi" junto a AG-008
    safe_set(ws, 4, 5, "Resultados en psi")

    print("AG-008 Omega…")
    r008 = write_table(ws, 5, "AG-008 Omega 1-24842", AG008, "$D$4")

    print("AG-034 PV350 (faltaba)…")
    r034 = write_table(ws, 31, "AG-034 Fluke PV350 MMI-CC-P-0012-2025", AG034, "$D$26")

    print("AG-052 Druck…")
    r052 = write_table(ws, 43, "AG-052 Druck 1-23207", AG052, "$D$20")

    update_meta(ws)
    patch_calculos(wb, r008, r034, r052)
    wb["Portada"]["D46"] = "=Calculos!N12"

    safe_set(
        ws,
        56,
        5,
        "Fuente: PDFs cert. EMP±=|sesgo|+U. Calculos!N12 elige tabla (≤350 AG-034 | ≤750 AG-052 | else AG-008).",
    )
    ws.cell(56, 5).font = Font(italic=True, size=8, color="666666")

    try:
        wb.save(out)
    except PermissionError:
        out = ALT
        wb.save(out)

    print("Guardado:", out)
    print(f"  AG-008: F{r008[0]}:I{r008[1]}")
    print(f"  AG-034: F{r034[0]}:I{r034[1]}  ← completo")
    print(f"  AG-052: F{r052[0]}:I{r052[1]}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
