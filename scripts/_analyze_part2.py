import openpyxl
from pathlib import Path

path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
lines = []

# Patrones row 5 labels E-I and rows 29-35
ws = wb["Patrones"]
lines.append("=== Patrones E5:I5 and F17 label row ===")
for r in [5, 17, 29, 30, 31, 32, 33, 34, 35]:
    parts = []
    for c in range(5, 10):
        v = ws.cell(r, c).value
        L = openpyxl.utils.get_column_letter(c)
        parts.append(f"{L}{r}={v!r}")
    lines.append(" | ".join(parts))

# Patrones full row 29-35 A-I
lines.append("\n=== Patrones rows 29-35 A-I ===")
for r in range(29, 36):
    parts = []
    for c in range(1, 10):
        cell = ws.cell(r, c)
        L = openpyxl.utils.get_column_letter(c)
        v = cell.value
        if isinstance(v, str) and v.startswith("="):
            parts.append(f"{L}{r}:{v}")
        else:
            parts.append(f"{L}{r}:{v!r}")
    lines.append("R"+str(r)+"| "+" | ".join(parts))

# Patrones catalog A62:J80 sample
lines.append("\n=== Patrones A62:J65 (catalog header area) ===")
for r in range(60, 66):
    parts = []
    for c in range(1, 11):
        v = ws.cell(r, c).value
        L = openpyxl.utils.get_column_letter(c)
        parts.append(f"{L}{r}:{v!r}")
    lines.append(" | ".join(parts))

# Calculos row 24-28 headers O-Y
ws = wb["Calculos"]
lines.append("\n=== Calculos rows 24-28 cols N-Y ===")
for r in range(24, 29):
    parts = []
    for c in range(14, 26):  # N=14 Y=25
        v = ws.cell(r, c).value
        L = openpyxl.utils.get_column_letter(c)
        if v is not None:
            parts.append(f"{L}{r}:{v!r}")
    if parts:
        lines.append(f"R{r}: "+" | ".join(parts))

lines.append("\n=== Calculos O28, P28, Q28, R28, S28, T28, U28, V28, W28, X28, Y28 ===")
for c in range(15, 26):
    L = openpyxl.utils.get_column_letter(c)
    v = ws.cell(28, c).value
    lines.append(f"{L}28: {v!r}")

# Search VLOOKUP on Patrones with col index 4 (EMP column I)
lines.append("\n=== Any formula referencing Patrones col I or VLOOKUP col 4 ===")
for sn in wb.sheetnames:
    ws = wb[sn]
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "Patrones!" in v:
                if "I$" in v or ",4," in v or "EMP" in v.upper():
                    lines.append(f"{sn}!{cell.coordinate}: {v}")

# Resultados references to Patrones?
lines.append("\n=== Resultados Patrones refs ===")
ws = wb["Resultados"]
for row in ws.iter_rows():
    for cell in row:
        v = cell.value
        if isinstance(v, str) and ("Patrones" in v or "Calculos!" in v and any(x in v for x in ["O28","Y28","V28"])):
            if "Patrones" in v:
                lines.append(f"Resultados!{cell.coordinate}: {v}")

wb.close()
Path("scripts/_excel_analysis_part2.txt").write_text("\n".join(lines), encoding="utf-8")
print(len(lines))
