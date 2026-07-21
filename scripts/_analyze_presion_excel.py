import openpyxl
from pathlib import Path
import json

out_lines = []
path = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato master auto Presion_SIN_REF.xlsm")
cert_path = Path("scripts/_cert_summary.json")

if cert_path.exists():
    out_lines.append("=== scripts/_cert_summary.json ===")
    out_lines.append(cert_path.read_text(encoding="utf-8"))
else:
    out_lines.append("scripts/_cert_summary.json does not exist")

if not path.exists():
    out_lines.append(f"FILE NOT FOUND: {path}")
else:
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    out_lines.append("Sheet names: " + ", ".join(wb.sheetnames))
    out_lines.append("")
    out_lines.append("=== 1. Patrones sheet rows 1-40, cols A-I ===")
    ws = wb["Patrones"]
    for r in range(1, 41):
        row_parts = []
        for c in range(1, 10):
            cell = ws.cell(row=r, column=c)
            col_letter = openpyxl.utils.get_column_letter(c)
            ref = f"{col_letter}{r}"
            if cell.value is None:
                val = ""
            elif isinstance(cell.value, str) and cell.value.startswith("="):
                val = cell.value
            else:
                val = repr(cell.value)
            row_parts.append(f"{ref}:{val}")
        out_lines.append(f"R{r}| " + " | ".join(row_parts))

    out_lines.append("")
    out_lines.append("=== 2. All formulas containing Patrones! ===")
    patron_refs = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "Patrones!" in v:
                    line = f"{sheet_name}!{cell.coordinate}: {v}"
                    out_lines.append(line)
                    patron_refs.append((sheet_name, cell.coordinate, v))

    out_lines.append("")
    out_lines.append(f"Total Patrones! references: {len(patron_refs)}")

    out_lines.append("")
    out_lines.append("=== 3. Calculos - key cells (Nominal/Patron/Incertidumbre/EMP) ===")
    if "Calculos" in wb.sheetnames:
        ws = wb["Calculos"]
        # dump first 50 rows A-Z for context
        for r in range(1, 51):
            for c in range(1, 27):
                cell = ws.cell(row=r, column=c)
                v = cell.value
                if v is None:
                    continue
                vs = str(v)
                if any(k in vs.upper() for k in ["PATRON", "NOMINAL", "INCERT", "EMP", "VLOOKUP", "INDEX", "MATCH", "Patrones"]):
                    col_letter = openpyxl.utils.get_column_letter(c)
                    out_lines.append(f"Calculos!{col_letter}{r}: {v}")

    out_lines.append("")
    out_lines.append("=== 4. VLOOKUP/INDEX/MATCH on EMP (all sheets) ===")
    emp_keywords = ["EMP", "VLOOKUP", "INDEX", "MATCH"]
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                vu = v.upper()
                if "EMP" in vu and ("VLOOKUP" in vu or "INDEX" in vu or "MATCH" in vu):
                    out_lines.append(f"{sheet_name}!{cell.coordinate}: {v}")
                elif "Patrones!" in v and ("EMP" in vu or "INCERT" in vu):
                    out_lines.append(f"{sheet_name}!{cell.coordinate}: {v}")

    # Patrones header row analysis
    out_lines.append("")
    out_lines.append("=== Patrones row 1 headers (A through I and beyond to M) ===")
    ws = wb["Patrones"]
    for c in range(1, 14):
        cell = ws.cell(row=1, column=c)
        col_letter = openpyxl.utils.get_column_letter(c)
        out_lines.append(f"{col_letter}1: {cell.value!r}")

    wb.close()

Path("scripts/_excel_analysis_output.txt").write_text("\n".join(out_lines), encoding="utf-8")
print("Wrote", len(out_lines), "lines")
