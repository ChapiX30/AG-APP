# -*- coding: utf-8 -*-
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

MM = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato Micrometro Exteriores mm .xlsx")
wb = openpyxl.load_workbook(MM, data_only=False)
ws = wb["Calculos"]

print("=== Merges around cert/header ===")
for m in ws.merged_cells.ranges:
    if m.min_row <= 15:
        print(" ", m)

print("\n=== Rows 24-40 A-B and key formulas ===")
for r in range(24, 40):
    parts = []
    for c in range(1, 12):
        v = ws.cell(r, c).value
        if v is not None:
            parts.append(f"{get_column_letter(c)}={repr(v)[:70]}")
    print(f"R{r}:", " | ".join(parts))

print("\n=== Helper X/AA and AC/AF headers ===")
for r in range(1, 6):
    parts = []
    for c in range(24, 33):
        v = ws.cell(r, c).value
        if v is not None:
            parts.append(f"{get_column_letter(c)}{r}={repr(v)[:50]}")
    if parts:
        print(" | ".join(parts))

print("\n=== Portada cert ===")
port = wb["Portada"]
for addr in ["J9", "J10", "H9", "E54"]:
    print(addr, repr(port[addr].value)[:100])

print("\n=== Pattern IDs ===")
pat = wb["Patrones"]
for r in range(1, 170):
    a, c, d = pat.cell(r, 1).value, pat.cell(r, 3).value, pat.cell(r, 4).value
    if a and "Instrumento" in str(a):
        print(f"R{r}: B={pat.cell(r,2).value} D={d}")
    if c and str(c).strip().upper() in ("ID:", "ID"):
        print(f"  ID row {r}: {d}")

print("\n=== Tecnico cell ===")
for r in range(15, 25):
    for c in range(15, 25):
        v = ws.cell(r, c).value
        if v and ("Calibr" in str(v) or "Angel" in str(v) or "Abraham" in str(v)):
            print(get_column_letter(c) + str(r), repr(v)[:60])

# formula count with J11
print("\n=== Formulas mentioning J11 or mm ===")
n = 0
for r in range(1, min(ws.max_row, 80) + 1):
    for c in range(1, min(ws.max_column, 32) + 1):
        v = ws.cell(r, c).value
        if isinstance(v, str) and v.startswith("=") and ("J11" in v or '"mm"' in v or "$X$" in v or "$AC$" in v):
            print(f"  {get_column_letter(c)}{r}: {v[:120]}")
            n += 1
            if n > 40:
                break
    if n > 40:
        break

wb.close()
