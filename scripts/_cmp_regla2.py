# -*- coding: utf-8 -*-
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter

MM = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato regla-flex mm.xlsx")
IN = Path(r"C:\Users\AG\Desktop\FORMATOS AG\Formato regla-flex in.xlsx")

wb = openpyxl.load_workbook(MM, data_only=False)
ws = wb["Calculos"]
print("=== Full M18/O18 MM ===")
print("M18:", ws["M18"].value)
print("O18:", ws["O18"].value)
print("M19:", ws["M19"].value)
print("T18:", ws["T18"].value)

print("\n=== Merges cert ===")
for m in ws.merged_cells.ranges:
    if m.min_row <= 8:
        print(" ", m)

print("\n=== Patrones MM ===")
pat = wb["Patrones"]
for r in range(1, 31):
    row = []
    for c in range(1, 9):
        v = pat.cell(r, c).value
        if v is not None:
            row.append(f"{get_column_letter(c)}{r}={repr(v)[:40]}")
    if row:
        print(" | ".join(row))

print("\n=== Portada ===")
p = wb["PORTADA"]
for addr in ["J9", "H9", "E35", "B5"]:
    print(addr, repr(p[addr].value)[:90] if p[addr].value else None)

# technician
print("\n=== Calibro area ===")
for r in range(7, 12):
    for c in range(12, 18):
        v = ws.cell(r, c).value
        if v is not None:
            print(f"{get_column_letter(c)}{r}={repr(v)[:50]}")

wb.close()

wb2 = openpyxl.load_workbook(IN, data_only=False)
pat2 = wb2["Patrones"]
print("\n=== Patrones IN (wider) ===")
for r in range(1, 31):
    row = []
    for c in range(1, 17):
        v = pat2.cell(r, c).value
        if v is not None:
            row.append(f"{get_column_letter(c)}{r}={repr(v)[:35]}")
    if row:
        print(" | ".join(row)[:200])
print("\nIN O18:", wb2["Calculos"]["O18"].value)
wb2.close()
